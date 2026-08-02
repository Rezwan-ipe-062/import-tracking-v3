"""Create the editable Country Thresholds.xlsx store for the Phase-3 urgency
layer. The file lives next to the raw sources in excel files\. If it is
missing or a cell is blank, rule_engine falls back to its built-in defaults.
"""
import sys
from pathlib import Path

SCRIPTS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import rule_engine

OUT = SCRIPTS_DIR.parent / "excel files" / "Country Thresholds.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Country Thresholds"

headers = ["Route Group", "LC - Urgent (days)", "LC - Critical (days)",
           "ETD - Urgent (days)", "ETD - Critical (days)"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True)

rows_data = [
    ("India", 45, 30, 30, 20),
    ("ASEAN", 60, 45, 40, 30),
    ("China/East Asia", 75, 60, 55, 45),
    ("Europe/Long-haul", 120, 90, 90, 75),
]
for i, r in enumerate(rows_data, 2):
    for c, v in enumerate(r, 1):
        ws.cell(row=i, column=c, value=v)
for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 22

ws2 = wb.create_sheet("How to use")
notes = [
    "Import Visibility Master - Country Thresholds (Phase 3)",
    "",
    "These values override the built-in starter thresholds in rule_engine.py.",
    "The urgency layer reads this file when it exists; if it is missing or a",
    "cell is blank, the engine falls back to the built-in defaults.",
    "",
    "Columns:",
    "- Route Group: India | ASEAN | China/East Asia | Europe/Long-haul",
    "- LC - Urgent/Critical (days): when LC Date is missing, how many days",
    "  before RDD makes the row Urgent / Critical.",
    "- ETD - Urgent/Critical (days): same for a missing ETD (schedule).",
    "",
    "Route groups map from the Eagle Eye From country code:",
    "  IN -> India",
    "  TH, SG -> ASEAN",
    "  CN, KR -> China/East Asia",
    "  DE, IT, CH, NL, GB, US -> Europe/Long-haul",
    "",
    "Starter values from the agreed business table (table.md). To finalise,",
    "validate with Nayeem Bhai / Logistics / Order Management, then edit this",
    "sheet and re-run clean_merge.py to regenerate the master.",
    "",
    "Safeguards baked into the engine:",
    "1. A blank field is never Critical just because it is blank.",
    "2. Source country is only used after Product Master confirmation - until",
    "   then the route comes from EE From (shipment origin).",
    '3. Final Docs / OBL "after ATD" rules are dropped (ATD blank in SAP).',
]
for i, line in enumerate(notes, 1):
    ws2.cell(row=i, column=1, value=line)
ws2.column_dimensions["A"].width = 110

wb.save(OUT)
print("written:", OUT)

th = rule_engine.load_country_thresholds(OUT)
print("LC India  (urgent, critical):", th["LC"]["India"])
print("LC ChinaEA:", th["LC"]["ChinaEA"])
print("ETD Europe:", th["ETD"]["Europe"])
