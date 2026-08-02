"""Single-file cleaner for the BD Tracker import file.

Run this script in VS Code (Run button / F5):

    1. A file dialog opens - pick the BD Tracker .xlsx file.
    2. The script reads the "Tracker file" sheet, keeps only the required
       fields, splits the partial-shipment suffix off the PO number, and
       normalises every PO to a 10-digit digit string so it can be joined
       with the other import files later (the merge step does the joining).
    3. It writes <name>_cleaned.xlsx next to the source file (the source
       file is never modified) and prints a summary to the console.

Kept rows stay at partial-shipment / line level - no PO rollup is done on
purpose. Every row for a PO stays a separate row.

Date columns are written as real Excel dates (yyyy-mm-dd), never as text, so
the output can be used for downstream date analysis. AGI is written as text so
codes like 40952 never become 40952.0.

The number of data rows is never hard-coded: rows are read lazily and the
scan stops after a long run of empty rows (with a peek-ahead safety net), so
it adapts whether the data ends at row 120, 150 or 1000 and never reaches the
data-validation dropdown lists Excel hides far below the table.

Requires: openpyxl  (pip install openpyxl)
"""

import datetime
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REQUIRED = {
    "overallstatus",
    "po",
    "agi",
    "lcdate",
    "sishareddate",
    "rdd",
    "etd",
    "eta",
    "obleblrcvddate",
    "finaldocsrcvddate",
}

OUT_COLUMNS = [
    "Overall Status",
    "PO",
    "Partial Shipment No.",
    "AGI",
    "LC Date",
    "SI Shared Date",
    "RDD",
    "ETD",
    "ETA",
    "OBL/EBL rcvd Date",
    "Final Docs rcvd Date",
]

DATE_KEYS = [
    "lcdate",
    "sishareddate",
    "rdd",
    "etd",
    "eta",
    "obleblrcvddate",
    "finaldocsrcvddate",
]

# Matches the dropdown list used in the tracker (Bank/Status columns) plus
# statuses observed in the data. Used only to flag unexpected values.
KNOWN_STATUSES = {
    "Completed",
    "LC yet to receive",
    "SI shared - Waiting for schedule & draft",
    "Schedule received - Waiting for Draft",
    "Draft received - waiting for OBL",
    "OBL Received - waiting for other docs",
    "Docs created, under Prasanna validation",
    "Hard copy pending",
    "Full set received",
    "HSBC Discrepancy / Approval pending",
}

DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%b-%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%Y.%m.%d",
)


def norm(s):
    return "".join(c for c in str(s).lower() if c.isalnum())


def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()


def is_blankish(v):
    if v is None:
        return True
    if isinstance(v, (int, float)) and v == 0:
        return True
    return str(v).strip().lower() in ("", "-", "--", "na", "n/a", "n.a.",
                                      "nan", "null", "none")


def find_header_row(sheet):
    for r in range(1, min(sheet.max_row, 6) + 1):
        names = set()
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(row=r, column=c).value
            if v is not None:
                names.add(norm(v))
        if {"overallstatus", "po", "agi"} <= names:
            return r
    return None


def build_column_map(sheet, header_row):
    colmap = {}
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=header_row, column=c).value
        if v is not None:
            colmap[norm(v)] = c
    missing = REQUIRED - set(colmap)
    return colmap, missing


def live_row(a, c, e):
    return not (is_blankish(a) and is_blankish(c) and is_blankish(e))


def iter_indexed_rows(ws, colmap, header_row):
    """Yield (excel_row_number, (A, C, E)) for the data region.

    Lazy scan with a dead-run break (50 consecutive empty rows) and a
    peek-ahead safety net (200 rows), so the table extent is discovered, not
    hard-coded, and the data-validation dropdown rows far below the table are
    never reached.
    """
    ca, cc, ce = (colmap["overallstatus"], colmap["po"], colmap["agi"])
    r = header_row + 1
    dead_run = 0
    while r <= ws.max_row:
        a = ws.cell(row=r, column=ca).value
        c = ws.cell(row=r, column=cc).value
        e = ws.cell(row=r, column=ce).value
        if live_row(a, c, e):
            dead_run = 0
            yield r, (a, c, e)
            r += 1
            continue
        dead_run += 1
        r += 1
        if dead_run < 50:
            continue
        # Suspected end - peek ahead before giving up.
        resumed = False
        for rr in range(r, min(r + 200, ws.max_row + 1)):
            a2 = ws.cell(row=rr, column=ca).value
            c2 = ws.cell(row=rr, column=cc).value
            e2 = ws.cell(row=rr, column=ce).value
            if live_row(a2, c2, e2):
                resumed = True
                break
        if not resumed:
            break
        dead_run = 0


def po_text(v):
    """Return a PO cell as plain digit text.

    Excel often stores PO numbers as numbers, and when a file is re-saved or
    edited the values silently become floats (6590028230.0). Strip the .0 so
    the 10-digit match works instead of dropping the row.
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return clean_str(v)


def parse_po(raw):
    """Return (base_10_digit_po, partial_shipment_no).

    Accepts '6590028256 - 2' -> ('6590028256', '2'),
    '6590028230' -> ('6590028230', '').
    Unparsable values come back as (raw, '') so they can be flagged.
    """
    s = po_text(raw)
    m = re.match(r"^(\d{10})\s*-\s*(\d+)$", s)
    if m:
        return m.group(1), m.group(2)
    m = re.match(r"^(\d{10})$", s)
    if m:
        return m.group(1), ""
    return s, ""


def parse_date_value(v):
    """Return (datetime_or_None, error_or_None). Time is stripped."""
    if v is None:
        return None, None
    if isinstance(v, datetime.datetime):
        return v.replace(hour=0, minute=0, second=0, microsecond=0), None
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day), None
    if isinstance(v, (int, float)):
        if v == 0:
            return None, None
        try:
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
            return d.replace(hour=0, minute=0, second=0, microsecond=0), None
        except (OverflowError, ValueError, TypeError):
            return None, "unreadable date value %r" % (v,)
    s = str(v).strip()
    if is_blankish(v):
        return None, None
    for fmt in DATE_FORMATS:
        try:
            d = datetime.datetime.strptime(s, fmt)
            return d.replace(hour=0, minute=0, second=0, microsecond=0), None
        except ValueError:
            continue
    return None, "invalid date: %r" % s


def as_agi_text(v):
    """AGI as text: '74211' not '74211.0' and never a float."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def cell_value(fcell, vcell):
    """Prefer the raw (formula) cell; fall back to the computed value."""
    if isinstance(fcell, str) and fcell.startswith("="):
        return vcell if vcell is not None else fcell
    return fcell


def clean_to_rows(src_path):
    """Clean the BD Tracker file in memory and return (headers, rows, info).

    ``rows`` is the list of kept output rows (same shape as OUT_COLUMNS).
    ``info`` carries the diagnostics needed for the console summary and is
    also used by clean_merge.py.
    """
    wb_formulas = openpyxl.load_workbook(src_path, data_only=False)
    wb_values = openpyxl.load_workbook(src_path, data_only=True)
    sheet_name = ("Tracker file" if "Tracker file" in wb_formulas.sheetnames
                  else wb_formulas.active.title)
    ws = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]

    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError('Could not find a header row containing "Overall Status" / "PO" / "AGI".')
    colmap, missing = build_column_map(ws, header_row)
    if missing:
        found = [ws.cell(row=header_row, column=c).value
                 for c in range(1, ws.max_column + 1)]
        raise ValueError("Missing expected columns: %s\nHeader found: %s"
                         % (", ".join(sorted(missing)), found))

    kept = []
    dropped = []
    review = []
    raw_pos = []
    n_blank_rows = 0
    n_formula_dates = 0
    kept_pos = []

    for r, (a, c, e) in iter_indexed_rows(ws, colmap, header_row):
        if is_blankish(a) and is_blankish(c) and is_blankish(e):
            n_blank_rows += 1
            continue

        status = clean_str(cell_value(ws.cell(row=r, column=colmap["overallstatus"]).value,
                                      ws_v.cell(row=r, column=colmap["overallstatus"]).value))
        po_raw = clean_str(cell_value(ws.cell(row=r, column=colmap["po"]).value,
                                      ws_v.cell(row=r, column=colmap["po"]).value))
        agi = as_agi_text(cell_value(ws.cell(row=r, column=colmap["agi"]).value,
                                     ws_v.cell(row=r, column=colmap["agi"]).value))

        base_po, partial = parse_po(po_raw)

        if base_po == "":
            dropped.append([r, po_raw, agi, "Blank / unparsable PO number"])
            continue
        if base_po == po_raw and not re.match(r"^\d{10}$", po_raw):
            dropped.append([r, po_raw, agi, "Unparsable PO format"])
            continue

        flags = []
        if status == "":
            flags.append("Blank Overall Status")
        elif status not in KNOWN_STATUSES:
            flags.append("Unrecognised Overall Status: %r" % status)
        if agi == "":
            flags.append("Blank AGI")

        out_row = [status, base_po, partial, agi]
        for key in DATE_KEYS:
            fc = ws.cell(row=r, column=colmap[key]).value
            vc = ws_v.cell(row=r, column=colmap[key]).value
            raw = cell_value(fc, vc)
            if isinstance(fc, str) and fc.startswith("="):
                n_formula_dates += 1
                flags.append("Formula in %s preserved" % key)
            dt, err = parse_date_value(raw)
            if err:
                flags.append("%s: %s" % (key, err))
            out_row.append(dt)

        kept.append(out_row)
        review.append([r, base_po, agi, "; ".join(flags)])
        kept_pos.append(base_po)
        raw_pos.append(po_raw)

    info = {
        "src_path": src_path,
        "sheet_name": sheet_name,
        "kept": kept,
        "review": review,
        "dropped": dropped,
        "kept_pos": kept_pos,
        "raw_pos": raw_pos,
        "n_blank_rows": n_blank_rows,
        "n_formula_dates": n_formula_dates,
    }
    return OUT_COLUMNS, kept, info


def write_rows(dst_path, headers, rows, date_cols=(), title="Cleaned Output",
               width=24, date_format="yyyy-mm-dd"):
    """Write headers + rows to a fresh workbook with basic styling."""
    out = openpyxl.Workbook()
    ws_out = out.active
    ws_out.title = title
    for c, name in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True)
    for i, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_out.cell(row=i, column=c, value=val)
            if c in date_cols and isinstance(val, datetime.datetime):
                cell.number_format = date_format
    for c in range(1, len(headers) + 1):
        ws_out.column_dimensions[ws_out.cell(row=1, column=c).column_letter].width = width
    ws_out.freeze_panes = "A2"
    out.save(dst_path)


def clean_workbook(src_path, dst_path):
    headers, kept, info = clean_to_rows(src_path)

    date_cols = set(range(5, 5 + len(DATE_KEYS)))  # cols 5..11
    write_rows(dst_path, headers, kept, date_cols=date_cols,
               title="Cleaned BD Tracker", width=24)

    dropped = info["dropped"]
    review = info["review"]
    kept_pos = info["kept_pos"]
    n_kept = len(kept)
    n_dropped = len(dropped)
    n_distinct = len(set(kept_pos))
    n_flagged = sum(1 for row in review if row[3] != "")
    print("=" * 60)
    print("BD Tracker cleaner - summary")
    print("=" * 60)
    print("Source          : %s" % src_path)
    print("Sheet           : %s" % info["sheet_name"])
    print("Kept            : %d rows across %d distinct PO numbers" % (n_kept, n_distinct))
    print("Dropped         : %d  (blank PO: %d, unparsable: %d)"
          % (n_dropped,
             sum(1 for d in dropped if "Blank" in d[3]),
             sum(1 for d in dropped if "Unparsable" in d[3])))
    print("Blank rows skip : %d" % info["n_blank_rows"])
    print("Rows flagged    : %d" % n_flagged)
    print("Formula dates   : %d (carried through)" % info["n_formula_dates"])
    print("Output          : %s" % dst_path)
    print("=" * 60)
    if n_flagged:
        print("Flagged rows to review (Source Row):")
        for row in review:
            if row[3] != "":
                print("  row %s PO=%s AGI=%s -> %s" % (row[0], row[1], row[2], row[3]))
    if n_dropped:
        print("Dropped rows (Source Row):")
        for d in dropped:
            print("  row %s PO=%r AGI=%s -> %s" % (d[0], d[1], d[2], d[3]))


def pick_file(title):
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    root.destroy()
    if path:
        return Path(path)
    p = input("No file chosen. Type the full path to the .xlsx: ").strip()
    return Path(p) if p else None


def main():
    src = pick_file("Select the BD Tracker .xlsx file")
    if src is None:
        print("No file selected - nothing to do. Exiting.")
        return
    if not src.exists():
        print("File not found: %s" % src)
        return
    dst = src.with_name(src.stem + "_cleaned.xlsx")
    try:
        clean_workbook(src, dst)
    except ValueError as exc:
        print("Error:", exc)


if __name__ == "__main__":
    main()
