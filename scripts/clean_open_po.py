"""Single-file cleaner for the Open PO extract.

Run this script in VS Code (Run button / F5):

    1. A file dialog opens - pick the Open PO .xlsx file.
    2. The script reads the "Data" sheet, keeps only the required fields,
       drops 62-series (local tolling) POs, drops blank PO numbers,
       splits the supplier/plant code from its name, and flags rows that
       need review.
    3. It writes <name>_cleaned.xlsx next to the source file (the source
       file is never modified) and prints a summary to the console.

Kept rows stay at PO-material line level. Duplicate PO numbers are kept on
purpose - one PO can have many material lines. Formula cells found in kept
columns are carried through to the output unchanged.

Requires: openpyxl  (pip install openpyxl)
"""

import datetime
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REQUIRED = {
    "material",
    "shorttext",
    "purchasingdocument",
    "suppliersupplyingplant",
    "stilltobedeliveredqty",
    "orderunit",
}

OUT_COLUMNS = [
    "Material (AGI)",
    "Short Text",
    "Purchasing Document",
    "Supplier Code",
    "Supplier Name",
    "Still to be Delivered (Qty)",
    "Order Unit",
]


def norm(s):
    return "".join(c for c in str(s).lower() if c.isalnum())


def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()


def agi_text(v):
    """Material/AGI as plain digits without leading zeros.

    The Open PO file stores the material as '0074211' but every other file
    (BD Tracker, Eagle Eye) uses '74211'. Normalise so the values join cleanly
    across files.
    """
    s = clean_str(v)
    s = s.lstrip("0")
    return s if s else "0"


def po_text(v):
    """Return a PO cell as plain digit text.

    Excel often stores PO numbers as numbers, and after re-saves or edits the
    values silently become floats (6590028230.0). Strip the .0 so the output
    stays clean digit text and the BD Tracker cleaner can match it.
    """
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return clean_str(v)


def find_header_row(sheet):
    for r in range(1, min(sheet.max_row, 6) + 1):
        names = set()
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(row=r, column=c).value
            if v is not None:
                names.add(norm(v))
        if "purchasingdocument" in names:
            return r
    return None


def build_column_map(sheet, header_row):
    colmap = {}
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=header_row, column=c).value
        if v is not None:
            colmap[norm(v)] = c
    missing = REQUIRED - set(colmap)
    return colmap, missing


def to_number(v):
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip().replace(",", "")
    if s == "":
        return None, "blank"
    try:
        return float(s), None
    except ValueError:
        return None, "non-numeric"


def parse_date(v):
    if v is None:
        return None, None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat(), None
    if isinstance(v, datetime.date):
        return v.isoformat(), None
    if isinstance(v, (int, float)):
        try:
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
            return d.date().isoformat(), None
        except (OverflowError, ValueError, TypeError):
            return None, "unreadable"
    s = str(v).strip()
    if s == "":
        return None, None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y",
                "%m/%d/%Y", "%d-%b-%Y", "%d-%m-%Y", "%Y.%m.%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat(), None
        except ValueError:
            continue
    return None, "invalid date: %r" % s


def split_supplier(full):
    if not full:
        return "", ""
    m = re.match(r"^(\d+)", full)
    if m:
        return m.group(1), full[m.end():].strip()
    return "", full


def clean_to_rows(src_path):
    """Clean the Open PO file in memory and return (headers, rows, info).

    ``rows`` is the list of kept output rows (same shape as OUT_COLUMNS).
    ``info`` carries the diagnostics needed for the console summary and is
    also used by clean_merge.py.
    """
    wb_formulas = openpyxl.load_workbook(src_path, data_only=False)
    wb_values = openpyxl.load_workbook(src_path, data_only=True)
    sheet_name = "Data" if "Data" in wb_formulas.sheetnames else wb_formulas.active.title
    ws = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]

    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError('Could not find a header row containing "Purchasing Document".')
    colmap, missing = build_column_map(ws, header_row)
    if missing:
        found = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
        raise ValueError("Missing expected columns: %s\nHeader found: %s"
                         % (", ".join(sorted(missing)), found))

    kept = []
    review = []
    dropped = []
    raw_pos = []
    raw_materials = []
    n_blank_lines = 0
    n_formulas = 0
    po_numbers = []

    for r in range(header_row + 1, ws.max_row + 1):
        raw = {}
        for key in REQUIRED:
            fcell = ws.cell(row=r, column=colmap[key])
            vcell = ws_v.cell(row=r, column=colmap[key])
            raw_val = fcell.value
            if isinstance(raw_val, str) and raw_val.startswith("="):
                eff = vcell.value if vcell.value is not None else raw_val
                raw[key] = (raw_val, eff, True)
            else:
                raw[key] = (raw_val, raw_val, False)

        if all(raw[k][0] in (None, "") for k in REQUIRED):
            n_blank_lines += 1
            continue

        material = agi_text(raw["material"][1])
        shorttext = clean_str(raw["shorttext"][1])
        po = po_text(raw["purchasingdocument"][1])
        supplier_full = clean_str(raw["suppliersupplyingplant"][1])
        unit = clean_str(raw["orderunit"][1])

        if po == "":
            dropped.append([r, material, shorttext, po, supplier_full, "Blank PO number"])
            continue
        if po.startswith("62"):
            dropped.append([r, material, shorttext, po, supplier_full,
                            "Out of scope - 62-series (local tolling)"])
            continue

        flags = []
        if not po.startswith("65"):
            flags.append("Unrecognised PO series - verify with Nayeem")

        if material == "":
            flags.append("Blank Material (AGI)")
        if shorttext == "":
            flags.append("Blank Short Text")
        if supplier_full == "":
            flags.append("Blank supplier")

        sup_code, sup_name = split_supplier(supplier_full)

        qty_write = None
        if raw["stilltobedeliveredqty"][2]:
            qty_write = raw["stilltobedeliveredqty"][0]
            flags.append("Formula in open qty preserved - not validated")
            n_formulas += 1
        else:
            qv = raw["stilltobedeliveredqty"][1]
            if qv in (None, ""):
                flags.append("Blank open qty")
            else:
                num, err = to_number(qv)
                if err == "non-numeric":
                    qty_write = clean_str(qv)
                    flags.append("Non-numeric open qty: %s" % clean_str(qv))
                elif num < 0:
                    qty_write = num
                    flags.append("Negative open qty")
                elif num == 0:
                    qty_write = num
                    flags.append("Zero open qty - confirm with Nayeem")
                else:
                    qty_write = num

        if unit == "":
            flags.append("Blank Order Unit")

        po_numbers.append(po)
        flags_str = "; ".join(flags) if flags else ""
        kept.append([
            material, shorttext, po, sup_code, sup_name,
            qty_write, unit,
        ])
        review.append([r, po, material, flags_str])
        raw_pos.append(raw["purchasingdocument"][1])
        raw_materials.append(raw["material"][1])

    info = {
        "src_path": src_path,
        "sheet_name": sheet_name,
        "kept": kept,
        "review": review,
        "dropped": dropped,
        "po_numbers": po_numbers,
        "raw_pos": raw_pos,
        "raw_materials": raw_materials,
        "n_blank_lines": n_blank_lines,
        "n_formulas": n_formulas,
    }
    return OUT_COLUMNS, kept, info


def write_rows(dst_path, headers, rows, date_cols=(), title="Cleaned Open PO",
               width=26, date_format="yyyy-mm-dd"):
    """Write headers + rows to a fresh workbook with basic styling."""
    out = openpyxl.Workbook()
    ws_out = out.active
    ws_out.title = title
    for c, name in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True)
    for i, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_out.cell(row=i, column=c, value=val)
            if c in date_cols and isinstance(val, datetime.datetime):
                cell.number_format = date_format
    for c in range(1, len(headers) + 1):
        ws_out.column_dimensions[ws_out.cell(row=1, column=c).column_letter].width = width
    ws_out.freeze_panes = "A2"
    out.save(dst_path)


def clean_workbook(src_path, dst_path):
    headers, kept, info = clean_to_rows(src_path)

    write_rows(dst_path, headers, kept)

    dropped = info["dropped"]
    review = info["review"]
    po_numbers = info["po_numbers"]
    n_kept = len(kept)
    n_dropped = len(dropped)
    n_distinct = len(set(po_numbers))
    n_flagged = sum(1 for row in review if row[3] != "")
    print("=" * 60)
    print("Open PO cleaner - summary")
    print("=" * 60)
    print("Source : %s" % src_path)
    print("Sheet  : %s" % info["sheet_name"])
    print("Kept   : %d lines across %d distinct PO numbers" % (n_kept, n_distinct))
    print("Dropped: %d  (62-series: %d, blank PO: %d)"
          % (n_dropped,
             sum(1 for d in dropped if "62-series" in d[5]),
             sum(1 for d in dropped if "Blank PO" in d[5])))
    print("Blank rows skipped : %d" % info["n_blank_lines"])
    print("Rows flagged       : %d" % n_flagged)
    print("Formulas preserved : %d" % info["n_formulas"])
    print("Output : %s" % dst_path)
    print("=" * 60)
    if n_flagged:
        print("Flagged rows to review (Source Row):")
        for row in review:
            if row[3] != "":
                print("  row %s [%s] %s -> %s" % (row[0], row[1], row[2], row[3]))
    if n_dropped:
        print("Dropped rows:")
        for d in dropped:
            print("  row %s %s -> %s" % (d[0], d[2], d[5]))


def pick_file():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title="Select the Open PO .xlsx file",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    root.destroy()
    if path:
        return Path(path)
    p = input("No file chosen. Type the full path to the .xlsx: ").strip()
    return Path(p) if p else None


def main():
    src = pick_file()
    if src is None:
        print("No file selected - nothing to do. Exiting.")
        return
    if not src.exists():
        print("File not found: %s" % src)
        return
    dst = src.with_name(src.stem + "_cleaned.xlsx")
    try:
        clean_workbook(src, dst)
    except ValueError as exc:
        print("Error:", exc)


if __name__ == "__main__":
    main()
