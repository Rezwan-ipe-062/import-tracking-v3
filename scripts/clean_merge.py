"""Build the Import Visibility Master by merging the three import files.

Run this script in VS Code (Run button / F5):

    1. Four file dialogs open in order - pick the raw Open PO, BD Tracker,
       Eagle Eye *and* the Country Thresholds .xlsx files. The thresholds file
       drives the Phase-3 urgency rules per route group (blank/missing cells
       fall back to built-in defaults); if you cancel, defaults are used.
    2. Each file is cleaned in memory by calling the matching cleaner's
       ``clean_to_rows()`` - no temporary files, no duplicate dialogs, and
       exactly the same cleaning rules as the standalone scripts.
    3. One single master sheet is written: one row per (Open PO material
       line x BD partial shipment), with the Open PO block, the BD Tracker
       block and the Eagle Eye block side by side. Eagle Eye containers are
       joined into cells ("; "-separated) because the master grain is the
       BD partial, not the container.
    4. It writes Import Visibility Master_merged.xlsx next to the Open PO
       source file (source files are never modified) and prints a summary.
       The workbook also contains Exceptions, Unmatched BD, Unmatched EE,
       Cleaning Log and Freshness sheets so data-quality issues stay with the
       deliverable instead of living only in the console.

Join key is the PO number only (10-digit text) - never AGI. Rows from BD and
Eagle Eye whose PO is absent from the Open PO extract are excluded from the
active master but are preserved in the Unmatched BD / Unmatched EE sheets.

Open quantity is shown ONCE per PO-material line and is labelled
non-additive, so summing it across partial-expanded rows can never
double-count. No BD row is ever silently discarded: every BD row for a PO is
kept, and multi-row partial groups are reported on the Exceptions sheet.

Requires: openpyxl  (pip install openpyxl)
"""

import datetime
import os
import tkinter as tk
from pathlib import Path
from tkinter import filedialog

import openpyxl
from openpyxl.styles import Font

import clean_bd_tracker
import clean_eagle_eye
import clean_open_po
import rule_engine

# ------------------------------------------------------------------ release
# Version stamp for this pipeline. Bump whenever the column layout, merge
# logic or rule thresholds change in a way that changes output. Written to the
# deliverable's a "Release" sheet / core properties so a given workbook can
# always be traced back to the exact code and thresholds that produced it.
PIPELINE_VERSION = "3.1.0"

# Column layout of the merged output (single master sheet).
# One row per (Open PO material line x BD partial shipment).
MERGE_COLUMNS = [
    # Open PO block
    "Material (AGI)",
    "Short Text",
    "Purchasing Document",
    "Supplier Code",
    "Supplier Name",
    "Still to be Delivered (Qty)",
    "Order Unit",
    # BD Tracker block (per partial-shipment row)
    "Partial Shipment No.",
    "Overall Status",
    "LC Date",
    "SI Shared Date",
    "RDD",
    "ETD",
    "BD Tracker ETA",
    "OBL/EBL rcvd Date",
    "Final Docs rcvd Date",
    # Eagle Eye block (joined into cells per PO)
    "From",
    "Container No.",
    "Tracking",
    "Status",
    "EE ETA",
    "EE ETD",
    # Phase-3 urgency layer (rule_engine.py)
    "Import Country",
    "Current EE Stage",
    "Container Assigned?",
    "Urgency",
    "Primary Reason",
    # Data-quality population layer
    "Population Status",
]

# Indexes of the derived (non-source) columns in a merged row.
COL_POPULATION = len(MERGE_COLUMNS) - 1
COL_URGENCY = MERGE_COLUMNS.index("Urgency")
COL_REASON = MERGE_COLUMNS.index("Primary Reason")

# Column indices (0-based) that hold Excel dates in each merged row.
MERGE_DATE_COLS = {9, 10, 11, 12, 13, 14, 15, 20, 21}

# Index of the Purchasing Document column inside an Open PO cleaned row.
OP_PO_IDX = 2
# Index of the base PO inside a BD Tracker cleaned row.
BD_PO_IDX = 1
# Index of the partial-shipment number inside a BD Tracker cleaned row.
BD_PARTIAL_IDX = 2
# Index of the cleansed DDPO inside an Eagle Eye cleaned row.
EE_PO_IDX = 2

# Control-sheet layouts.
EXCEPTIONS_COLUMNS = ["Source File", "Source Sheet", "Source Row", "Raw PO",
                      "Cleaned PO", "Issue / Flag", "Refresh Date"]
UNMATCHED_BD_COLUMNS = ["PO", "AGI", "Overall Status", "Source Sheet",
                        "Source Row", "Reason"]
UNMATCHED_EE_COLUMNS = ["PO", "Container No.", "Tracking", "Status", "ETA",
                        "Source Sheet", "Source Row", "Reason"]
CLEANING_LOG_COLUMNS = ["Source File", "Source Sheet", "Kept Rows",
                        "Dropped Rows", "Flagged Rows", "Distinct POs",
                        "Cleaned At"]
FRESHNESS_COLUMNS = ["Source File", "Sheet", "File Modified", "Master Refresh",
                     "Stale (> 3 days)?"]


def join_unique(values):
    """Join non-empty values with '; ', keeping first-seen order."""
    seen = set()
    out = []
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s == "" or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return "; ".join(out)


def join_positional(values):
    """Join values with '; ', one slot per source row (no dedup).

    Keeps list lengths equal across columns so container/tracking/status stay
    positionally aligned. Empty values become blank (the source convention
    would put '-'; blank is cleaner for downstream analysis).
    """
    parts = []
    for v in values:
        if v is None or str(v).strip() == "":
            parts.append("")
        else:
            parts.append(str(v).strip())
    return "; ".join(parts)


def earliest(values):
    """Earliest datetime from a list of datetimes (None entries ignored)."""
    vals = [d for d in values if d is not None]
    return min(vals) if vals else None


def first_non_blank(values):
    for v in values:
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return ""


def _rule_columns(bd, ees, thresholds=None, status=None, open_qty=None):
    """Build the Phase-3 urgency columns for one merged row (rule_engine.py).

    ``bd`` is the cleaned BD row (or None); ``ees`` the PO's Eagle Eye rows.
    ``thresholds`` is an optional rules dict (see
    :func:`rule_engine.load_country_thresholds`); when None the engine uses
    its built-in defaults. Returns the 5 values: Import Country, Current EE
    Stage, Container Assigned?, Urgency, Primary Reason.
    """
    from_code = first_non_blank([e[0] for e in ees]) if ees else ""
    import_country = rule_engine.route_of(from_code) or ""

    stage = rule_engine.current_ee_stage([e[5] for e in ees]) if ees else None
    stage_text = "" if stage is None else str(stage)

    if not ees:
        container_flag = ""
    else:
        container_flag = "Yes" if any(
            e[3] is not None and str(e[3]).strip() not in ("", "-") for e in ees
        ) else "No"

    if bd is None:
        rdd = lc = etd_bd = eta_bd = None
    else:
        rdd, lc = bd[6], bd[4]
        etd_bd, eta_bd = bd[7], bd[8]
    etd_ee = earliest([e[7] for e in ees]) if ees else None
    eta_ee = earliest([e[6] for e in ees]) if ees else None

    urgency, reason = rule_engine.classify(
        rdd=rdd, lc=lc, etd_bd=etd_bd, etd_ee=etd_ee,
        eta_bd=eta_bd, eta_ee=eta_ee, from_code=from_code,
        thresholds=thresholds, status=status, open_qty=open_qty,
    )
    return [import_country, stage_text, container_flag, urgency, reason]


def group_by(rows, key_idx):
    groups = {}
    for row in rows:
        groups.setdefault(row[key_idx], []).append(row)
    return groups


def find_agi_mismatches(op_rows, bd_rows):
    """Return (bd_index, base_po, bd_agi, op_materials) for BD rows whose AGI
    is not among that PO's Open PO material lines.

    The merge joins on PO number only, so a BD row pointing at a different
    AGI would otherwise attach to the Open PO row silently. AGI is compared
    after stripping leading zeros ('059769' matches '59769'). Only POs that
    exist in the Open PO extract are checked.
    """
    op_mats = {}
    for r in op_rows:
        op_mats.setdefault(r[OP_PO_IDX], set()).add(r[0].lstrip("0"))
    out = []
    for i, b in enumerate(bd_rows):
        mats = op_mats.get(b[BD_PO_IDX])
        if mats is None:
            continue
        agi = str(b[3]).lstrip("0")
        if agi and agi not in mats:
            out.append((i, b[BD_PO_IDX], b[3], sorted(mats)))
    return out


def merge_rows(openpo_rows, bd_rows, ee_rows,
               op_info=None, bd_info=None, ee_info=None,
               thresholds=None):
    """Merge cleaned rows into master rows.

    ``thresholds`` is an optional per-route rules dict (from
    ``rule_engine.load_country_thresholds``); when None the engine's built-in
    defaults are used.

    Returns (merged_rows, summary_counts). Row shape follows MERGE_COLUMNS.
    """
    bd_by_po = group_by(bd_rows, BD_PO_IDX)
    ee_by_po = group_by(ee_rows, EE_PO_IDX)

    merged = []
    multi_partial_pos = set()
    duplicate_partials = []
    seen_qty = set()
    not_bd = 0
    not_ee = 0
    not_both = 0
    severity_counts = {"Critical": 0, "Urgent": 0, "Monitor": 0, "Data Review": 0}
    population_counts = {"Active": 0, "Quantity Review": 0}

    for op in openpo_rows:
        po = op[OP_PO_IDX]
        bds = bd_by_po.get(po, [])
        ees = ee_by_po.get(po, [])

        # Group this PO's BD rows by partial-shipment number.
        partial_groups = {}
        for b in bds:
            partial_groups.setdefault(b[BD_PARTIAL_IDX], []).append(b)

        # A PO without any BD rows still gets one row (blank BD block).
        if partial_groups:
            partials = [k for k in partial_groups if k] + [k for k in partial_groups if not k]
        else:
            partials = [""]

        non_empty_partials = [k for k in partial_groups if k]
        if len(non_empty_partials) > 1:
            multi_partial_pos.add(po)

        qty_key = (po, op[0])

        for p in partials:
            group = partial_groups.get(p) or [None]
            if len(group) > 1:
                duplicate_partials.append((po, p, len(group)))
            for bd in group:
                row = list(op)
                # Open quantity is non-additive: show it once per PO-material.
                if qty_key in seen_qty:
                    row[5] = ""
                else:
                    seen_qty.add(qty_key)
                if bd is None:
                    row += [""] * 9
                else:
                    row += [bd[2], bd[0], bd[4], bd[5],
                            bd[6], bd[7], bd[8], bd[9], bd[10]]
                if ees:
                    row += [
                        join_unique([e[0] for e in ees]),
                        join_positional([e[3] for e in ees]),
                        join_positional([e[4] for e in ees]),
                        join_positional([e[5] for e in ees]),
                        earliest([e[6] for e in ees]),
                        earliest([e[7] for e in ees]),
                    ]
                else:
                    row += [""] * 6
                row += _rule_columns(bd, ees, thresholds,
                                     status=bd[0] if bd else None,
                                     open_qty=op[5])
                row += [rule_engine.population_status(op[5])]
                if not bds:
                    not_bd += 1
                    if not ees:
                        not_both += 1
                if not ees:
                    not_ee += 1
                sev = row[COL_URGENCY]
                population_counts[row[COL_POPULATION]] = \
                    population_counts.get(row[COL_POPULATION], 0) + 1
                if row[COL_POPULATION] == "Active":
                    severity_counts[sev] = severity_counts.get(sev, 0) + 1
                merged.append(row)

    summary = {
        "rows": len(merged),
        "po_count": len({r[OP_PO_IDX] for r in merged}),
        "extra_rows": len(merged) - len(openpo_rows),
        "multi_partial_pos": sorted(multi_partial_pos),
        "duplicate_partials": duplicate_partials,
        "not_bd": not_bd,
        "not_ee": not_ee,
        "not_both": not_both,
        "severity": severity_counts,
        "population": population_counts,
    }
    return merged, summary


def build_control_sheets(op_info, bd_info, ee_info,
                         op_rows, bd_rows, ee_rows, src_paths):
    """Return a dict of {sheet_title: (headers, rows, date_cols)}.

    Builds the Exceptions, Unmatched BD, Unmatched EE, Cleaning Log and
    Freshness sheets from the cleaner infos so data-quality issues stay in the
    deliverable. ``src_paths`` is the ordered list of the three raw files.
    """
    refresh = datetime.datetime.now().replace(microsecond=0)
    op_pos = set(op_info["po_numbers"])

    # ---------------------------------------------------------- Exceptions
    exceptions = []

    def exc(source, sheet, src_row, raw_po, cleaned_po, issue):
        exceptions.append([source, sheet, src_row, raw_po, cleaned_po,
                           issue, refresh])

    for d in op_info["dropped"]:
        exc("Open PO", op_info["sheet_name"], d[0], d[3], "", d[5])
    for r in op_info["review"]:
        if r[3]:
            exc("Open PO", op_info["sheet_name"], r[0], r[1], r[1], r[3])
    for d in bd_info["dropped"]:
        exc("BD Tracker", bd_info["sheet_name"], d[0], d[1], "", d[3])
    for r in bd_info["review"]:
        if r[3]:
            exc("BD Tracker", bd_info["sheet_name"], r[0], r[1], r[1], r[3])
    for d in ee_info["dropped"]:
        exc("Eagle Eye", ee_info["sheet_name"], d[0], d[1], "", d[2])
    for r in ee_info["review"]:
        if r[2]:
            exc("Eagle Eye", ee_info["sheet_name"], r[0], r[1], r[1],
                "; ".join(r[2]))

    # AGI cross-check: BD rows whose AGI is missing from the PO's Open PO
    # material lines (join is by PO only, so the mismatch is silent unless
    # flagged here). Master data is left untouched.
    for i, base_po, bd_agi, op_mats in find_agi_mismatches(op_rows, bd_rows):
        exc("BD Tracker", bd_info["sheet_name"],
            bd_info["review"][i][0] if bd_info else "",
            bd_info["raw_pos"][i] if bd_info else base_po,
            base_po,
            "AGI mismatch: BD AGI %s not in OP material(s) %s"
            % (bd_agi, ", ".join(op_mats)))

    # Multi-material POs that also have BD rows: the BD block is duplicated
    # across every material line (the grain is material x partial), so flag
    # the duplication instead of letting it pass silently.
    op_mats = {}
    for op_i, r in enumerate(op_rows):
        op_mats.setdefault(r[OP_PO_IDX], [op_i, set()])
        op_mats[r[OP_PO_IDX]][1].add(r[0])
    bd_has = set(b[BD_PO_IDX] for b in bd_rows)
    for po, (op_i, mats) in sorted(op_mats.items()):
        if len(mats) > 1 and po in bd_has:
            exc("Open PO", op_info["sheet_name"],
                op_info["review"][op_i][0] if op_info else "",
                po, po,
                "Multi-material PO with BD rows - BD block duplicated "
                "across %d material lines" % len(mats))

    # ------------------------------------------------- Unmatched BD / EE
    unmatched_bd = []
    for i, b in enumerate(bd_rows):
        if b[BD_PO_IDX] not in op_pos:
            src_row = bd_info["review"][i][0] if bd_info else ""
            unmatched_bd.append([b[BD_PO_IDX], b[3], b[0],
                                 bd_info["sheet_name"] if bd_info else "",
                                 src_row, "PO not in Open PO extract"])

    unmatched_ee = []
    for i, e in enumerate(ee_rows):
        if e[EE_PO_IDX] not in op_pos:
            src_row = ee_info["review"][i][0] if ee_info else ""
            unmatched_ee.append([e[EE_PO_IDX], e[3], e[4], e[5], e[6],
                                 ee_info["sheet_name"] if ee_info else "",
                                 src_row, "PO not in Open PO extract"])

    # ------------------------------------------- Cleaning Log + Freshness
    cleaning_log = []
    freshness = []
    source_defs = [
        ("Open PO", op_info, op_rows, src_paths[0]),
        ("BD Tracker", bd_info, bd_rows, src_paths[1]),
        ("Eagle Eye", ee_info, ee_rows, src_paths[2]),
    ]
    for label, info, rows, path in source_defs:
        sheet = info["sheet_name"] if info else ""
        if label == "Open PO":
            distinct = len(set(info["po_numbers"]))
            n_flagged = sum(1 for r in info["review"] if r[3])
        else:
            distinct = len(set(info["kept_pos"]))
            n_flagged = sum(1 for r in info["review"]
                            if (r[3] if len(r) > 3 else bool(r[2])))
        cleaning_log.append([os.path.basename(str(path)), sheet, len(rows),
                             len(info["dropped"]), n_flagged, distinct, refresh])
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(str(path)))
        stale = "Yes" if (refresh - mtime).days > 3 else ""
        freshness.append([os.path.basename(str(path)), sheet, mtime, refresh,
                          stale])

    return {
        "Exceptions": (EXCEPTIONS_COLUMNS, exceptions, {6}),
        "Unmatched BD": (UNMATCHED_BD_COLUMNS, unmatched_bd, set()),
        "Unmatched EE": (UNMATCHED_EE_COLUMNS, unmatched_ee, {4}),
        "Cleaning Log": (CLEANING_LOG_COLUMNS, cleaning_log, {6}),
        "Freshness": (FRESHNESS_COLUMNS, freshness, {2, 3}),
    }


RELEASE_COLUMNS = ["Field", "Value"]
RECONCILIATION_COLUMNS = ["PO", "AGI", "Material", "Overall Status",
                          "Open Qty", "Urgency", "Requires", "Reason"]


def build_release_sheet(summary, refresh, thresholds, thresholds_src):
    """Return (headers, rows, date_cols) for the Release sheet.

    Records the pipeline version, refresh timestamp, thresholds source and the
    headline counts so a workbook can be audited back to the code+config that
    produced it. ``thresholds_src`` is a readable label (path suffix or
    "built-in defaults").
    """
    sev = summary["severity"]
    pop = summary.get("population", {})
    rows = [
        ["Pipeline", "Import Visibility Master - Phase 3"],
        ["Version", PIPELINE_VERSION],
        ["Refreshed", refresh],
        ["Thresholds source", thresholds_src],
        ["Threshold routes", str(len((thresholds or {}) or
                                     rule_engine.DEFAULT_THRESHOLDS.get("LC", {})))],
        ["Master rows", summary["rows"]],
        ["POs", summary["po_count"]],
        ["Not in BD", summary["not_bd"]],
        ["Not in EE", summary["not_ee"]],
        ["Neither", summary["not_both"]],
        ["Critical", sev.get("Critical", 0)],
        ["Urgent", sev.get("Urgent", 0)],
        ["Monitor", sev.get("Monitor", 0)],
        ["Data Review", sev.get("Data Review", 0)],
        ["Population Active", pop.get("Active", 0)],
        ["Population Quantity Review", pop.get("Quantity Review", 0)],
    ]
    return RELEASE_COLUMNS, rows, {2}


def build_reconciliation(merged, summary):
    """Return (headers, rows, date_cols) for a Reconciliation queue.

    Lists the rows that need a human decision and are excluded from risk
    totals: Quantity Review population rows, and any Data Review row that is
    caused by an urgency/status conflict (rather than plain RDD-missing or
    route-unknown), plus Completed-with-open-quantity conflicts.
    """
    rows = []
    for r in merged:
        pop = r[COL_POPULATION]
        urg = r[COL_URGENCY]
        reason = r[COL_REASON]
        if pop == "Quantity Review":
            rows.append([r[OP_PO_IDX], r[0], r[0], r[8],
                         r[5], urg, "Quantity Review - confirm open qty",
                         reason])
        elif "complete" in (reason or "").lower() or "open quantity" in \
                (reason or "").lower():
            rows.append([r[OP_PO_IDX], r[0], r[0], r[8], r[5], urg,
                         "Status vs open conflict",
                         reason])
    return RECONCILIATION_COLUMNS, rows, set()


# ---------------------------------------------------------------------------
# Workbook writer - single Import Visibility Master + control sheets.
# ---------------------------------------------------------------------------


def write_workbook(dst_path, sheets):
    """Write an ordered dict of {sheet_title: (headers, rows, date_cols)}.

    The first sheet becomes the active sheet; every sheet gets bold headers,
    yyyy-mm-dd number formats on its date columns and a frozen header row.
    """
    out = openpyxl.Workbook()
    out.properties.title = "Import Visibility Master - Phase 3 v%s" % PIPELINE_VERSION
    out.properties.description = "Regenerated %s (see Release sheet)" \
        % datetime.datetime.now()
    out.properties.creator = "Planning Team Product Build V3"
    out.remove(out.active)
    for title, (headers, rows, date_cols) in sheets.items():
        ws = out.create_sheet(title=title)
        for c, name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=name)
            cell.font = Font(bold=True)
        for i, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=c, value=val)
                if c - 1 in date_cols and isinstance(val, datetime.datetime):
                    cell.number_format = "yyyy-mm-dd"
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 24
        ws.freeze_panes = "A2"
    out.active = out.worksheets[0]
    out.save(dst_path)


def pick_file(title):
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    root.destroy()
    if path:
        return Path(path)
    p = input("No file chosen. Type the full path to the .xlsx: ").strip()
    return Path(p) if p else None


def main():
    src1 = pick_file("Select the Open PO .xlsx file")
    if src1 is None:
        print("No file selected - nothing to do. Exiting.")
        return
    src2 = pick_file("Select the BD Tracker .xlsx file")
    if src2 is None:
        print("No file selected - nothing to do. Exiting.")
        return
    src3 = pick_file("Select the Eagle Eye .xlsx file")
    if src3 is None:
        print("No file selected - nothing to do. Exiting.")
        return
    src4 = pick_file("Select the Country Thresholds .xlsx file")

    for s in (src1, src2, src3):
        if s is None:
            print("No file selected - nothing to do. Exiting.")
            return
        if not s.exists():
            print("File not found: %s" % s)
            return

    # The Country Thresholds file is optional: if it is cancelled or does not
    # exist we fall back to the engine's built-in defaults (per-cell where
    # blank). Canceling it must NOT abort the merge.
    thresholds = None
    if src4 is not None and src4.exists():
        thresholds = rule_engine.load_country_thresholds(src4)
        if thresholds is None:
            print("Could not read country thresholds - using built-in defaults.")
        else:
            print("Country thresholds loaded from: %s" % src4)
    else:
        print("Country thresholds not provided - using built-in defaults.")

    print("Cleaning Open PO ...")
    _, openpo_rows, op_info = clean_open_po.clean_to_rows(src1)
    print("Cleaning BD Tracker ...")
    _, bd_rows, bd_info = clean_bd_tracker.clean_to_rows(src2)
    print("Cleaning Eagle Eye ...")
    _, ee_rows, ee_info = clean_eagle_eye.clean_to_rows(src3)

    op_pos = set(op_info["po_numbers"])
    refresh = datetime.datetime.now().replace(microsecond=0)

    threshold_label = (str(src4) if src4 is not None else None)
    if thresholds is None:
        threshold_src = "built-in defaults"
    else:
        threshold_src = threshold_label or "built-in defaults"

    merged, summary = merge_rows(openpo_rows, bd_rows, ee_rows,
                                 op_info, bd_info, ee_info,
                                 thresholds=thresholds)
    control = build_control_sheets(op_info, bd_info, ee_info,
                                   openpo_rows, bd_rows, ee_rows,
                                   [src1, src2, src3])
    sheets = {"Import Visibility Master": (MERGE_COLUMNS, merged,
                                           MERGE_DATE_COLS)}
    sheets["Release"] = build_release_sheet(summary, refresh, thresholds,
                                            threshold_src)
    sheets["Reconciliation"] = build_reconciliation(merged, summary)
    sheets.update(control)

    dst = src1.with_name("Import Visibility Master_merged.xlsx")
    write_workbook(dst, sheets)

    print("=" * 60)
    print("Import Visibility Master - merge summary")
    print("=" * 60)
    print("Open PO lines    : %d" % len(openpo_rows))
    print("BD Tracker rows  : %d" % len(bd_rows))
    print("Eagle Eye rows   : %d" % len(ee_rows))
    print("Master rows      : %d (%d POs, %d extra from partials)"
          % (summary["rows"], summary["po_count"], summary["extra_rows"]))
    print("Multi-partial POs: %s" % (", ".join(summary["multi_partial_pos"])
                                     or "-"))
    print("Not in BD (rows) : %d   Not in EE (rows): %d   Neither: %d"
          % (summary["not_bd"], summary["not_ee"], summary["not_both"]))
    sev = summary["severity"]
    print("Urgency (Active) : Critical %d | Urgent %d | Monitor %d | Data Review %d"
          % (sev.get("Critical", 0), sev.get("Urgent", 0),
             sev.get("Monitor", 0), sev.get("Data Review", 0)))
    pop = summary.get("population", {})
    print("Population       : Active %d | Quantity Review %d"
          % (pop.get("Active", 0), pop.get("Quantity Review", 0)))
    for name in sheets:
        print("Sheet            : %s" % name)
    print("Output           : %s" % dst)
    print("=" * 60)


if __name__ == "__main__":
    main()
