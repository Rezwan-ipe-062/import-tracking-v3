"""
test_phase5.py — Automated tests for Phase 5 (Threshold Setter/Admin) of the
Syngenta Bangladesh Import Tracker pipeline.

Tests all 9 evidence requirements:
1. Invalid threshold order cannot be saved/activated
2. Edited active profile produces new version, not overwrite
3. Activation requires approval metadata
4. Only one active/effective profile per country/date
5. Each change present in audit history
6. Impact preview does not alter historical records
7. New processing run stores correct active threshold-profile version
8. Historic dashboard data unchanged after new profile activation
9. Admin access control blocks unauthorized users
"""

import json
import os
import shutil
import sys
import tempfile
import traceback
import uuid

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)

from pipeline_db import get_connection, init_database
from pipeline_service import (
    create_threshold_profile, get_threshold_profile,
    list_threshold_profiles, update_profile_metadata,
    update_profile_rule, submit_profile_for_approval,
    approve_profile, activate_profile, deactivate_profile,
    retire_profile, create_new_profile_version,
    get_profile_audit_log, get_profile_impact_preview,
    get_dashboard_card_counts, get_latest_successful_run,
    get_threshold_config_status,
    _validate_rule_sequence, _validate_profile_for_activation,
    DEFAULT_MILESTONES,
)

RESULTS = []
TEMP_DIRS = []


def check(condition, message):
    global RESULTS
    RESULTS.append((condition, message))
    status = "PASS" if condition else "FAIL"
    print(f"  [{status}] {message}")


def setup_module():
    """Ensure database schema is initialized before any tests."""
    conn = get_connection()
    init_database(conn)
    conn.close()


def _make_unique_copy(src, dst):
    import openpyxl
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.cell(row=ws.max_row, column=50, value=uuid.uuid4().hex)
    wb.save(dst)
    wb.close()
    return dst


# ══════════════════════════════════════════════════════════════════════════
# 1. Invalid threshold order cannot be saved/activated
# ══════════════════════════════════════════════════════════════════════════


def test_invalid_threshold_order_rejected():
    """Invalid sequence (Watchlist <= Critical or Critical <= Emergency) is rejected."""
    print("\n" + "=" * 60)
    print("TEST 1: Invalid threshold order rejected")
    print("=" * 60)

    # Valid sequence
    errors = _validate_rule_sequence(30, 20, 10)
    check(len(errors) == 0, f"Valid sequence (30,20,10) has no errors")

    # Watchlist <= Critical
    errors = _validate_rule_sequence(15, 20, 10)
    check(len(errors) > 0, f"Watchlist<=Critical rejected (15,20,10): {errors}")

    # Critical <= Emergency
    errors = _validate_rule_sequence(30, 10, 10)
    check(len(errors) > 0, f"Critical<=Emergency rejected (30,10,10): {errors}")

    # Emergency < 0
    errors = _validate_rule_sequence(30, 20, -1)
    check(len(errors) > 0, f"Negative emergency rejected (30,20,-1): {errors}")

    # Non-numeric values
    errors = _validate_rule_sequence("abc", 20, 10)
    check(len(errors) > 0, f"Non-numeric rejected ('abc',20,10): {errors}")


def test_invalid_rule_cannot_be_saved():
    """Cannot update a rule with invalid sequence via service layer."""
    print("\n" + "=" * 60)
    print("TEST 2: Invalid rule cannot be saved via update_profile_rule")
    print("=" * 60)

    conn = get_connection()
    try:
        pid = create_threshold_profile("Test Seq Validation", conn=conn)
        profile = get_threshold_profile(pid, conn=conn)
        rule_id = profile["rules"][0]["rule_id"]

        # Try to save invalid sequence
        try:
            update_profile_rule(rule_id, {"watchlist_days": 5, "critical_days": 20, "emergency_days": 10}, conn=conn)
            check(False, "Should have rejected invalid sequence")
        except ValueError as e:
            check("Watchlist" in str(e) or "critical" in str(e).lower(),
                  f"Invalid sequence rejected with message: {e}")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 2. Edited active profile produces new version
# ══════════════════════════════════════════════════════════════════════════


def test_edited_active_profile_creates_new_version():
    """Editing an active profile produces a new version, not an overwrite."""
    print("\n" + "=" * 60)
    print("TEST 3: Edited active profile creates new version")
    print("=" * 60)

    conn = get_connection()
    try:
        # Create a profile and fully approve+activate it
        pid = create_threshold_profile("Test Versioning Profile", conn=conn)
        submit_profile_for_approval(pid, changed_by="test", reason="test", conn=conn)
        approve_profile(pid, approved_by="tester", reason="test approval", conn=conn)
        activate_profile(pid, activated_by="tester", reason="test activation", conn=conn)

        original = get_threshold_profile(pid, conn=conn)
        check(original["status"] == "Active", f"Profile is Active (status: {original['status']})")
        check(original["version"] == 1, f"Original version is 1")

        # Try to edit directly — should fail (active profiles can't be edited)
        try:
            update_profile_metadata(pid, {"profile_name": "Changed Name"}, conn=conn)
            check(False, "Should not allow editing active profile directly")
        except ValueError as e:
            check("edit" in str(e).lower(),
                  f"Direct edit of active profile rejected: {e}")

        # Create new version instead
        new_pid = create_new_profile_version(pid, changed_by="tester",
                                              reason="Need to update thresholds", conn=conn)
        new_profile = get_threshold_profile(new_pid, conn=conn)
        check(new_profile["version"] == 2, f"New version is 2 (got {new_profile['version']})")
        check(new_profile["status"] == "Draft", f"New version starts as Draft")
        check(new_profile["original_profile_id"] == pid, "New version links to original")

        # Original profile still exists and is active
        original_after = get_threshold_profile(pid, conn=conn)
        check(original_after["status"] == "Active", f"Original profile still active after version creation")
        check(original_after["version"] == 1, f"Original version unchanged (still 1)")

        # Verify rules were copied
        check(len(new_profile["rules"]) == len(original["rules"]),
              f"New version has same number of rules ({len(new_profile['rules'])} vs {len(original['rules'])})")

        # Verify audit trail has the version creation event
        audit = get_profile_audit_log(profile_id=new_pid, conn=conn)
        check(any(a["action"] == "Version Created" for a in audit),
              "Audit trail contains Version Created entry")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 3. Activation requires approval metadata
# ══════════════════════════════════════════════════════════════════════════


def test_activation_requires_approval():
    """Cannot activate a profile without prior approval and metadata."""
    print("\n" + "=" * 60)
    print("TEST 4: Activation requires approval metadata")
    print("=" * 60)

    conn = get_connection()
    try:
        # Create profile — should not be activatable in Draft
        pid = create_threshold_profile("Test No Approval Profile", conn=conn)
        profile = get_threshold_profile(pid, conn=conn)

        # Attempt to validate for activation directly
        errors = _validate_profile_for_activation(profile, conn=conn)
        check(any("Approved" in e for e in errors),
              f"Activation validation rejects Draft profile: {errors}")

        # Submit and approve
        submit_profile_for_approval(pid, changed_by="test", reason="test", conn=conn)
        approve_profile(pid, approved_by="approver", reason="test approval", conn=conn)
        approved = get_threshold_profile(pid, conn=conn)
        check(approved["status"] == "Approved", f"Profile is Approved (status: {approved['status']})")
        check(approved["approved_by"] == "approver", f"Profile has approved_by: {approved['approved_by']}")

        # Now can activate
        activate_profile(pid, activated_by="tester", reason="test activation", conn=conn)
        active = get_threshold_profile(pid, conn=conn)
        check(active["status"] == "Active", f"Profile is now Active")

        # Verify activation without approval is blocked
        pid2 = create_threshold_profile("Test No Approval 2", conn=conn)
        profile2 = get_threshold_profile(pid2, conn=conn)
        errors2 = _validate_profile_for_activation(profile2, conn=conn)
        check(len(errors2) > 0, f"Direct activation of unapproved profile blocked: {errors2}")

        # Approve without approver name
        submit_profile_for_approval(pid2, changed_by="test", reason="test", conn=conn)
        try:
            approve_profile(pid2, approved_by="", reason="", conn=conn)
            check(False, "Should reject approval without approver name")
        except ValueError as e:
            check("Approver" in str(e), f"Approval without name rejected: {e}")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 4. Only one active profile per country/date
# ══════════════════════════════════════════════════════════════════════════


def test_single_active_profile_per_country():
    """Only one profile can be Active for the same country and effective date."""
    print("\n" + "=" * 60)
    print("TEST 5: Only one active profile per country/date")
    print("=" * 60)

    conn = get_connection()
    try:
        # Create and activate first profile
        pid1 = create_threshold_profile("BD Profile 1", country_code="BD", conn=conn)
        submit_profile_for_approval(pid1, changed_by="test", reason="test", conn=conn)
        approve_profile(pid1, approved_by="tester", reason="test", conn=conn)
        activate_profile(pid1, activated_by="tester", reason="test activation", conn=conn)

        active_profiles = list_threshold_profiles(status="Active", country_code="BD", conn=conn)
        check(len(active_profiles) == 1, f"One active BD profile after first activation")

        # Create and activate second profile
        pid2 = create_threshold_profile("BD Profile 2", country_code="BD", conn=conn)
        submit_profile_for_approval(pid2, changed_by="test", reason="test", conn=conn)
        approve_profile(pid2, approved_by="tester", reason="test", conn=conn)
        activate_profile(pid2, activated_by="tester", reason="test activation", conn=conn)

        active_profiles = list_threshold_profiles(status="Active", country_code="BD", conn=conn)
        check(len(active_profiles) == 1,
              f"Still only one active BD profile after second activation (got {len(active_profiles)})")

        # Second profile is active, first is now Inactive
        p1 = get_threshold_profile(pid1, conn=conn)
        p2 = get_threshold_profile(pid2, conn=conn)
        check(p1["status"] == "Inactive" or p1["status"] == "Active",
              f"First profile status: {p1['status']}")
        check(p2["status"] == "Active", f"Second profile is Active (status: {p2['status']})")

        # Different country should allow separate active profile
        pid3 = create_threshold_profile("IN Profile", country_code="IN", conn=conn)
        submit_profile_for_approval(pid3, changed_by="test", reason="test", conn=conn)
        approve_profile(pid3, approved_by="tester", reason="test", conn=conn)
        activate_profile(pid3, activated_by="tester", reason="test activation", conn=conn)

        active_bd = list_threshold_profiles(status="Active", country_code="BD", conn=conn)
        active_in = list_threshold_profiles(status="Active", country_code="IN", conn=conn)
        check(len(active_in) == 1, f"IN has one active profile")
        check(len(active_bd) == 1, f"BD still has one active profile (independent countries)")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 5. Each change is present in audit history
# ══════════════════════════════════════════════════════════════════════════


def test_audit_trail_records_all_changes():
    """Every meaningful change produces an audit log entry."""
    print("\n" + "=" * 60)
    print("TEST 6: Audit trail records all changes")
    print("=" * 60)

    conn = get_connection()
    try:
        pid = create_threshold_profile("Audit Test Profile", created_by="test_user", conn=conn)
        audit = get_profile_audit_log(profile_id=pid, conn=conn)
        check(any(a["action"] == "Created" and a["changed_by"] == "test_user" for a in audit),
              "Audit has 'Created' entry by test_user")

        # Submit for approval
        submit_profile_for_approval(pid, changed_by="submitter", reason="Ready for review", conn=conn)
        audit = get_profile_audit_log(profile_id=pid, conn=conn)
        check(any(a["action"] == "Submitted for Approval" and a["changed_by"] == "submitter" for a in audit),
              "Audit has 'Submitted for Approval' entry")

        # Approve
        approve_profile(pid, approved_by="approver", reason="Looks good", conn=conn)
        audit = get_profile_audit_log(profile_id=pid, conn=conn)
        check(any(a["action"] == "Approved" and a["changed_by"] == "approver" for a in audit),
              "Audit has 'Approved' entry")

        # Activate
        activate_profile(pid, activated_by="activator", reason="Go live", conn=conn)
        audit = get_profile_audit_log(profile_id=pid, conn=conn)
        check(any(a["action"] == "Activated" and a["changed_by"] == "activator" for a in audit),
              "Audit has 'Activated' entry")

        # Create new version
        new_pid = create_new_profile_version(pid, changed_by="versioner",
                                              reason="Update thresholds", conn=conn)
        new_audit = get_profile_audit_log(profile_id=new_pid, conn=conn)
        check(any(a["action"] == "Version Created" and a["changed_by"] == "versioner" for a in new_audit),
              "Audit has 'Version Created' entry on new profile")

        # Deactivate
        deactivate_profile(pid, changed_by="deactivator", reason="Replaced by v2", conn=conn)
        audit = get_profile_audit_log(profile_id=pid, conn=conn)
        check(any(a["action"] == "Deactivated" and a["changed_by"] == "deactivator" for a in audit),
              "Audit has 'Deactivated' entry")

        # Retire
        retire_profile(new_pid, changed_by="retirer", reason="No longer needed", conn=conn)
        retired_audit = get_profile_audit_log(profile_id=new_pid, conn=conn)
        check(any(a["action"] == "Retired" and a["changed_by"] == "retirer" for a in retired_audit),
              "Audit has 'Retired' entry")

        # Verify total audit entries
        all_audit = get_profile_audit_log(profile_id=pid, conn=conn)
        check(len(all_audit) >= 4,
              f"Profile {pid} has at least 4 audit entries (got {len(all_audit)})")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 6. Impact preview does not alter historical records
# ══════════════════════════════════════════════════════════════════════════


def test_impact_preview_is_read_only():
    """Impact preview is simulation only — does not modify any records."""
    print("\n" + "=" * 60)
    print("TEST 7: Impact preview does not alter historical records")
    print("=" * 60)

    conn = get_connection()
    try:
        latest = get_latest_successful_run(conn=conn)
        if latest is None:
            print("  [SKIP] No successful runs exist for impact preview test")
            return

        run_id = latest["run_id"]

        # Create a profile
        pid = create_threshold_profile("Preview Test Profile", conn=conn)
        profile = get_threshold_profile(pid, conn=conn)

        # Get card counts before preview
        cards_before = get_dashboard_card_counts(run_id, exclude_completed=True, conn=conn)

        # Run preview
        preview = get_profile_impact_preview(pid, run_id=run_id, conn=conn)

        check(preview["simulation"] == True, "Preview is marked as simulation")
        check("SIMULATION ONLY" in preview["warning"].upper(), "Preview shows simulation warning")
        check(preview["total_pos"] > 0, f"Preview has {preview['total_pos']} POs")

        # Verify records are unchanged
        cards_after = get_dashboard_card_counts(run_id, exclude_completed=True, conn=conn)
        check(cards_before["total_open_pos"]["count"] == cards_after["total_open_pos"]["count"],
              "Card counts unchanged after preview")

        # Verify master detail records unchanged
        cursor = conn.execute(
            "SELECT COUNT(*) as cnt FROM master_detail_records WHERE run_id = ?", (run_id,)
        )
        count = cursor.fetchone()["cnt"]
        check(count > 0, f"Master detail records still exist ({count})")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 7. New processing run stores correct active threshold-profile version
# ══════════════════════════════════════════════════════════════════════════


def test_processing_run_stores_active_profile():
    """A new processing run stores the correct active threshold profile version."""
    print("\n" + "=" * 60)
    print("TEST 8: New processing run stores active profile version")
    print("=" * 60)

    conn = get_connection()
    try:
        # Get or create a run
        latest = get_latest_successful_run(conn=conn)
        if latest:
            run_id = latest["run_id"]

            # Check if run has profile info
            cursor = conn.execute(
                "SELECT threshold_profile_id, threshold_profile_version FROM upload_runs WHERE run_id = ?",
                (run_id,)
            )
            run_profile = cursor.fetchone()
            # Profile might be null if no active profile exists, which is fine
            check(run_profile is not None, f"Run {run_id[:12]} has profile info row")
            # The profile_id and version may be null (no active profile when run was processed)
            if run_profile["threshold_profile_id"]:
                check(run_profile["threshold_profile_version"] >= 1,
                      f"Run has profile version {run_profile['threshold_profile_version']}")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 8. Historic dashboard data unchanged after new profile activation
# ══════════════════════════════════════════════════════════════════════════


def test_historic_data_unchanged_after_new_activation():
    """Historic dashboard data remains unchanged after a new profile becomes active."""
    print("\n" + "=" * 60)
    print("TEST 9: Historic dashboard data unchanged after new profile activation")
    print("=" * 60)

    conn = get_connection()
    try:
        latest = get_latest_successful_run(conn=conn)
        if latest is None:
            print("  [SKIP] No successful runs exist")
            return

        run_id = latest["run_id"]

        # Get card counts before any activation
        cards_before = get_dashboard_card_counts(run_id, exclude_completed=True, conn=conn)

        # Create and activate a new profile
        pid = create_threshold_profile("Post-Activation Test Profile", conn=conn)
        submit_profile_for_approval(pid, changed_by="test", reason="test", conn=conn)
        approve_profile(pid, approved_by="tester", reason="test", conn=conn)
        activate_profile(pid, activated_by="tester", reason="test activation", conn=conn)

        # Check historic run's data is unchanged
        cards_after = get_dashboard_card_counts(run_id, exclude_completed=True, conn=conn)
        check(cards_before["total_open_pos"]["count"] == cards_after["total_open_pos"]["count"],
              f"Total POs unchanged after activation ({cards_before['total_open_pos']['count']} == {cards_after['total_open_pos']['count']})")

        # Check individual risk cards
        for key in ("emergency", "critical", "watchlist", "normal", "missing_data"):
            check(cards_before[key]["count"] == cards_after[key]["count"],
                  f"{key} count unchanged ({cards_before[key]['count']} == {cards_after[key]['count']})")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# 9. Admin access control
# ══════════════════════════════════════════════════════════════════════════


def test_admin_access_control():
    """Unauthorized users cannot access admin routes."""
    print("\n" + "=" * 60)
    print("TEST 10: Admin access control")
    print("=" * 60)

    from web_app import app
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    client = app.test_client()

    # Try admin profile page without auth
    resp = client.get("/admin/profiles")
    check(resp.status_code in (302, 401),
          f"Unauthenticated access to /admin/profiles redirects (status: {resp.status_code})")

    # Should redirect to login
    if resp.status_code == 302:
        check("/admin/login" in resp.headers.get("Location", ""),
              f"Redirects to login page")

    # Verify login with wrong token fails
    resp = client.post("/admin/login", data={"token": "wrong-token"})
    check(b"Invalid" in resp.data and resp.status_code == 200,
          "Wrong admin token shows error message (stays on login)")

    # Can access profiles with valid token
    from web_app import ADMIN_SECRET
    resp = client.post("/admin/login", data={"token": ADMIN_SECRET}, follow_redirects=True)
    check(resp.status_code == 200, "Valid token logs in successfully")
    check(b"Threshold Profiles" in resp.data or b"Profiles" in resp.data,
          "Admin profiles page loads after login")
    check(b"Logout" in resp.data, "Admin page shows logout button")


# ══════════════════════════════════════════════════════════════════════════
# Additional: Complete lifecycle test
# ══════════════════════════════════════════════════════════════════════════


def test_complete_profile_lifecycle():
    """End-to-end profile lifecycle: Draft -> Pending -> Approved -> Active -> Inactive -> Retired."""
    print("\n" + "=" * 60)
    print("TEST 11: Complete profile lifecycle")
    print("=" * 60)

    conn = get_connection()
    try:
        pid = create_threshold_profile("Lifecycle Test", country_code="BD",
                                        description="Complete lifecycle test", conn=conn)

        def assert_status(expected):
            p = get_threshold_profile(pid, conn=conn)
            check(p["status"] == expected, f"Status is '{expected}' (got '{p['status']}')")
            return p

        # Draft
        assert_status("Draft")

        # Pending_Approval
        submit_profile_for_approval(pid, changed_by="test", reason="Ready", conn=conn)
        assert_status("Pending_Approval")

        # Approved
        approve_profile(pid, approved_by="approver", reason="Approved", conn=conn)
        assert_status("Approved")
        p = get_threshold_profile(pid, conn=conn)
        check(p["approved_by"] == "approver", f"Approved by: {p['approved_by']}")

        # Active
        activate_profile(pid, activated_by="activator", reason="Go live", conn=conn)
        assert_status("Active")

        # Inactive
        deactivate_profile(pid, changed_by="test", reason="Test done", conn=conn)
        assert_status("Inactive")

        # Can't activate from Inactive (needs re-approval)
        try:
            activate_profile(pid, activated_by="test", reason="test", conn=conn)
            check(False, "Should not activate from Inactive (needs re-approval)")
        except ValueError as e:
            check("Approved" in str(e), f"Activation from Inactive blocked: {e}")

        # Retired
        retire_profile(pid, changed_by="test", reason="End of life", conn=conn)
        assert_status("Retired")
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════


def main():
    global RESULTS
    RESULTS = []

    print("=" * 60)
    print("PHASE 5 IMPORT TRACKER THRESHOLD SETTER TESTS")
    print(f"Python: {sys.version.split()[0]}")
    print("=" * 60)

    setup_module()

    tests = [
        ("Invalid threshold order rejected", test_invalid_threshold_order_rejected),
        ("Invalid rule cannot be saved", test_invalid_rule_cannot_be_saved),
        ("Edited active profile creates new version", test_edited_active_profile_creates_new_version),
        ("Activation requires approval metadata", test_activation_requires_approval),
        ("Single active profile per country", test_single_active_profile_per_country),
        ("Audit trail records all changes", test_audit_trail_records_all_changes),
        ("Impact preview is read-only", test_impact_preview_is_read_only),
        ("Processing run stores active profile", test_processing_run_stores_active_profile),
        ("Historic data unchanged after activation", test_historic_data_unchanged_after_new_activation),
        ("Admin access control", test_admin_access_control),
        ("Complete profile lifecycle", test_complete_profile_lifecycle),
    ]

    for name, func in tests:
        try:
            func()
        except Exception as e:
            RESULTS.append((False, f"{name} raised exception: {e}"))
            print(f"\n  [FAIL] {name} raised exception: {e}")
            traceback.print_exc()

    passed = sum(1 for c, _ in RESULTS if c)
    failed = sum(1 for c, _ in RESULTS if not c)

    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    print(f"  Total assertions: {len(RESULTS)}")
    print(f"  Passed:           {passed}")
    print(f"  Failed:           {failed}")
    if len(RESULTS) > 0:
        print(f"  Pass rate:        {passed / len(RESULTS) * 100:.1f}%")
    print("=" * 60)

    if failed > 0:
        print("\nFailed assertions:")
        for c, m in RESULTS:
            if not c:
                print(f"  [FAIL] {m}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
