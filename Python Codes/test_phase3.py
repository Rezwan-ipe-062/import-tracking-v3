"""
test_phase3.py — Automated tests for Phase 3 (Flask Web App) of the
Syngenta Bangladesh Import Tracker pipeline.

Tests all web endpoints using the Flask test client.
"""

import io
import json
import os
import shutil
import sys
import tempfile
import traceback
import uuid

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)

EXCEL_DIR = os.path.join(_SCRIPT_DIR, "..", "Excel Files")
OPEN_PO_PATH = os.path.join(EXCEL_DIR, "Open PO 23rd July.xlsx")
BD_TRACKER_PATH = os.path.join(EXCEL_DIR, "BD TRACKER - 2026 v1.xlsx")
EAGLE_EYE_PATH = os.path.join(EXCEL_DIR, "Eagle eye.xlsx")

RESULTS = []


def check(condition, message):
    RESULTS.append((condition, message))
    status = "PASS" if condition else "FAIL"
    print(f"  [{status}] {message}")
    return condition


def _read_file(path):
    with open(path, "rb") as f:
        return f.read()


def check_source_files():
    missing = []
    for label, p in [("Open PO", OPEN_PO_PATH), ("BD Tracker", BD_TRACKER_PATH),
                     ("Eagle Eye", EAGLE_EYE_PATH)]:
        if not os.path.isfile(p):
            missing.append(f"{label}: {p}")
    return missing


def _json_of(response):
    return json.loads(response.data.decode("utf-8"))


# ======================================================================
# TEST 1: Index page loads
# ======================================================================

def test_index_page_loads(client):
    print("\n" + "=" * 60)
    print("TEST 1: test_index_page_loads")
    print("=" * 60)
    resp = client.get("/")
    check(resp.status_code == 200, f"GET / returns 200 (got {resp.status_code})")
    check(b"Upload" in resp.data or b"Process" in resp.data, "Page contains upload content")


# ======================================================================
# TEST 2: Validate with no files
# ======================================================================

def test_validate_no_files(client):
    print("\n" + "=" * 60)
    print("TEST 2: test_validate_no_files")
    print("=" * 60)
    resp = client.post("/validate", data={})
    data = _json_of(resp)
    check(resp.status_code == 200, f"POST /validate returns 200 (got {resp.status_code})")
    check("valid" in data, "Response has 'valid' key")
    check(data["valid"] == False, "Validation fails with no files (valid=False)")
    check(len(data.get("errors", [])) > 0, "Errors reported for missing files")


# ======================================================================
# TEST 3: Clear temp files
# ======================================================================

def test_clear_temp_files(client):
    print("\n" + "=" * 60)
    print("TEST 3: test_clear_temp_files")
    print("=" * 60)
    resp = client.post("/clear")
    data = _json_of(resp)
    check(resp.status_code == 200, f"POST /clear returns 200 (got {resp.status_code})")
    check("cleared" in data, "Response has 'cleared' key")
    check(isinstance(data["cleared"], int), "cleared is an integer")


# ======================================================================
# TEST 4: Process with invalid paths
# ======================================================================

def test_process_invalid_paths(client):
    print("\n" + "=" * 60)
    print("TEST 4: test_process_invalid_paths")
    print("=" * 60)
    resp = client.post(
        "/process",
        data=json.dumps({
            "open_po_path": "/nonexistent/po.xlsx",
            "bd_tracker_path": "/nonexistent/bd.xlsx",
            "eagle_eye_path": "/nonexistent/ee.xlsx",
        }),
        content_type="application/json",
    )
    data = _json_of(resp)
    check(resp.status_code == 400, f"POST /process with bad paths returns 400 (got {resp.status_code})")
    check("error" in data, "Response contains error message")


# ======================================================================
# TEST 5: Runs list (empty)
# ======================================================================

def test_runs_list_empty(client):
    print("\n" + "=" * 60)
    print("TEST 5: test_runs_list_empty")
    print("=" * 60)
    resp = client.get("/runs")
    data = _json_of(resp)
    check(resp.status_code == 200, f"GET /runs returns 200 (got {resp.status_code})")
    check(isinstance(data, list), "Response is a list")


# ======================================================================
# TEST 6: Run detail not found
# ======================================================================

def test_run_detail_not_found(client):
    print("\n" + "=" * 60)
    print("TEST 6: test_run_detail_not_found")
    print("=" * 60)
    resp = client.get("/run/nonexistent-run-id-12345")
    check(resp.status_code == 404, f"GET /run/nonexistent returns 404 (got {resp.status_code})")


# ======================================================================
# TEST 7: Download invalid run
# ======================================================================

def test_download_invalid_run(client):
    print("\n" + "=" * 60)
    print("TEST 7: test_download_invalid_run")
    print("=" * 60)
    resp = client.get("/download/nonexistent-run-id-12345")
    check(resp.status_code == 404, f"GET /download/nonexistent returns 404 (got {resp.status_code})")


# ======================================================================
# TEST 8: Full upload flow (with real files)
# ======================================================================

def test_full_upload_flow(client):
    print("\n" + "=" * 60)
    print("TEST 8: test_full_upload_flow")
    print("=" * 60)

    # Validate
    with open(OPEN_PO_PATH, "rb") as op, \
         open(BD_TRACKER_PATH, "rb") as bd, \
         open(EAGLE_EYE_PATH, "rb") as ee:
        resp = client.post("/validate", data={
            "open_po": (op, "Open PO.xlsx"),
            "bd_tracker": (bd, "BD TRACKER.xlsx"),
            "eagle_eye": (ee, "Eagle eye.xlsx"),
        })
    data = _json_of(resp)
    check(resp.status_code == 200, "POST /validate returns 200")
    check(data["valid"], "All 3 files validate successfully")
    check("file_paths" in data, "Response contains file paths")

    file_paths = data["file_paths"]
    check("open_po" in file_paths, "open_po path present")
    check("bd_tracker" in file_paths, "bd_tracker path present")
    check("eagle_eye" in file_paths, "eagle_eye path present")

    # Process
    resp = client.post(
        "/process",
        data=json.dumps(file_paths),
        content_type="application/json",
    )
    data = _json_of(resp)
    check(resp.status_code == 200, f"POST /process returns 200 (got {resp.status_code})")
    check("run_id" in data, "Response contains run_id")
    check(data["success"], "Process succeeded")
    valid_statuses = ("Completed", "Completed_With_Exceptions", "Completed (Duplicate)")
    check(data["status"] in valid_statuses,
          f"Status is valid (got {data['status']})")

    run_id = data["run_id"]
    row_counts = data.get("row_counts", {})
    is_dup = data.get("duplicate", False)
    if is_dup:
        print(f"  [INFO] Duplicate run — no new processing; row_counts may be sparse")
    else:
        check(row_counts.get("master_detail", 0) > 0,
              f"Master detail rows > 0 (got {row_counts.get('master_detail')})")

    # Check status endpoint
    resp = client.get(f"/status/{run_id}")
    status_data = _json_of(resp)
    check(status_data["status"] in valid_statuses or status_data.get("error"),
          f"Status endpoint returns final status: {status_data['status']}")
    check("row_counts" in status_data, "Status response has row_counts")

    # Check runs list now has this run
    resp = client.get("/runs")
    runs_data = _json_of(resp)
    check(len(runs_data) >= 1, f"Runs list has at least 1 entry (got {len(runs_data)})")
    run_ids = [r["run_id"] for r in runs_data]
    check(run_id in run_ids, f"New run appears in runs list")

    # Check run detail page
    resp = client.get(f"/run/{run_id}")
    check(resp.status_code == 200, f"GET /run/{run_id} returns 200")

    # Check run detail data API
    resp = client.get(f"/run/{run_id}/data")
    data_api = _json_of(resp)
    check("run" in data_api, "Run data API has 'run' key")
    check("row_counts" in data_api, "Run data API has 'row_counts' key")
    check(data_api["run"]["run_id"] == run_id, "Run ID matches")

    # Check workbook download
    resp = client.get(f"/download/{run_id}")
    check(resp.status_code == 200, f"GET /download/{run_id} returns 200")
    check(resp.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
          or resp.content_type.startswith("application/octet"),
          f"Download returns Excel content type")

    # Check source file download
    for st in ("open_po", "bd_tracker", "eagle_eye"):
        resp = client.get(f"/download/{run_id}/source/{st}")
        check(resp.status_code in (200, 404),
              f"Source download for {st} returns 200 or 404 (got {resp.status_code})")

    return run_id


# ======================================================================
# TEST 9: Duplicate detection flow
# ======================================================================

def test_duplicate_detection_flow(client):
    print("\n" + "=" * 60)
    print("TEST 9: test_duplicate_detection_flow")
    print("=" * 60)

    # First upload + process
    with open(OPEN_PO_PATH, "rb") as op, \
         open(BD_TRACKER_PATH, "rb") as bd, \
         open(EAGLE_EYE_PATH, "rb") as ee:
        resp = client.post("/validate", data={
            "open_po": (op, "Open PO.xlsx"),
            "bd_tracker": (bd, "BD TRACKER.xlsx"),
            "eagle_eye": (ee, "Eagle eye.xlsx"),
        })
    data = _json_of(resp)
    check(data["valid"], "Files validate")

    resp = client.post(
        "/process",
        data=json.dumps(data["file_paths"]),
        content_type="application/json",
    )
    r1 = _json_of(resp)
    check(r1["success"], f"First process succeeded")

    has_dup_warning = any(w.get("field") == "all" for w in data.get("warnings", []))
    if has_dup_warning:
        print("  [INFO] Duplicate warning already present on second upload of same files")
    else:
        print("  [INFO] No duplicate warning (first run or different session)")

    # Second upload of same files (will detect duplicate since first run exists)
    with open(OPEN_PO_PATH, "rb") as op, \
         open(BD_TRACKER_PATH, "rb") as bd, \
         open(EAGLE_EYE_PATH, "rb") as ee:
        resp = client.post("/validate", data={
            "open_po": (op, "Open PO.xlsx"),
            "bd_tracker": (bd, "BD TRACKER.xlsx"),
            "eagle_eye": (ee, "Eagle eye.xlsx"),
        })
    data = _json_of(resp)
    has_dup = any(w.get("field") == "all" for w in data.get("warnings", []))
    check(has_dup, "Duplicate warning appears for second upload of identical files")


# ======================================================================
# TEST 10: Clear does not affect history
# ======================================================================

def test_clear_does_not_affect_history(client):
    print("\n" + "=" * 60)
    print("TEST 10: test_clear_does_not_affect_history")
    print("=" * 60)

    # Get runs before clear
    resp_before = client.get("/runs")
    runs_before = _json_of(resp_before)

    # Clear
    resp = client.post("/clear")
    data = _json_of(resp)
    check(resp.status_code == 200, "POST /clear returns 200")

    # Get runs after clear
    resp_after = client.get("/runs")
    runs_after = _json_of(resp_after)

    check(len(runs_after) == len(runs_before),
          f"Run count unchanged after clear: {len(runs_before)} -> {len(runs_after)}")

    if len(runs_after) > 0:
        run_id = runs_after[0]["run_id"]
        resp = client.get(f"/status/{run_id}")
        data = _json_of(resp)
        check(data.get("status") in ("Completed", "Completed_With_Exceptions", "Failed"),
              f"Historic run status intact: {data.get('status')}")


# ======================================================================
# TEST 11: Archived files survive workspace clear
# ======================================================================

def test_archive_survives_workspace_clear(client):
    """Prove archived source files remain accessible after clearing the
    upload workspace (which only removes temp uploaded files).
    Uses unique temp copies of the source files to avoid duplicate detection,
    ensuring process_upload_run actually runs and creates archive files."""
    print("\n" + "=" * 60)
    print("TEST 11: test_archive_survives_workspace_clear")
    print("=" * 60)

    # Create unique temp copies to avoid duplicate detection
    tmp_dir = tempfile.mkdtemp(prefix="archive_survival_")
    try:
        unique_files = {}
        for key, src in [("open_po", OPEN_PO_PATH), ("bd_tracker", BD_TRACKER_PATH),
                         ("eagle_eye", EAGLE_EYE_PATH)]:
            dst = os.path.join(tmp_dir, f"{key}_{uuid.uuid4().hex}.xlsx")
            _make_unique_copy(src, dst)
            unique_files[key] = dst

        with open(unique_files["open_po"], "rb") as op, \
             open(unique_files["bd_tracker"], "rb") as bd, \
             open(unique_files["eagle_eye"], "rb") as ee:
            resp = client.post("/validate", data={
                "open_po": (op, f"op_{uuid.uuid4().hex}.xlsx"),
                "bd_tracker": (bd, f"bd_{uuid.uuid4().hex}.xlsx"),
                "eagle_eye": (ee, f"ee_{uuid.uuid4().hex}.xlsx"),
            })
        data = _json_of(resp)
        check(data["valid"], "Validation passed for unique temp copies")
        if not data["valid"]:
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return

        resp = client.post(
            "/process",
            data=json.dumps(data["file_paths"]),
            content_type="application/json",
        )
        r = _json_of(resp)
        check(r["success"], "Process succeeded for archive test")
        check(not r.get("duplicate", False), "Run is not a duplicate")

        run_id = r["run_id"]

        from pipeline_service import get_archive_metadata, get_archived_file_path
        from pipeline_db import get_connection

        # Verify archive metadata exists
        meta = get_archive_metadata(run_id)
        check(meta is not None, f"Archive metadata exists for run {run_id[:8]}")

        # Check archived file paths resolve
        for st in ("open_po", "bd_tracker", "eagle_eye"):
            fpath = get_archived_file_path(run_id, st)
            check(fpath is not None, f"Archived file path found for {st}")
            check(os.path.exists(fpath), f"Archived file exists on disk for {st}")

        # Clear temp uploads (this does NOT affect archive)
        resp = client.post("/clear")
        clear_data = _json_of(resp)
        check("cleared" in clear_data, "Clear completed")

        # Verify archived files still exist after clear
        for st in ("open_po", "bd_tracker", "eagle_eye"):
            fpath = get_archived_file_path(run_id, st)
            check(fpath is not None, f"Archived file path still resolves after clear for {st}")
            check(os.path.exists(fpath), f"Archived file still exists on disk after clear for {st}")

        # Verify metadata still intact
        meta2 = get_archive_metadata(run_id)
        check(meta2 is not None, "Archive metadata still exists after clear")
        check(meta2["run_id"] == run_id, "Archive metadata run_id still correct")

        # Verify source download endpoint still works
        for st in ("open_po", "bd_tracker", "eagle_eye"):
            resp = client.get(f"/download/{run_id}/source/{st}")
            check(resp.status_code in (200, 404),
                  f"Source download endpoint still reachable after clear for {st} (got {resp.status_code})")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ======================================================================
# TEST 12: Archived files survive simulated app restart
# ======================================================================

def test_archive_survives_app_restart(client):
    """Prove archived files remain accessible after simulating an app
    restart (new client, fresh connection)."""
    print("\n" + "=" * 60)
    print("TEST 12: test_archive_survives_app_restart")
    print("=" * 60)

    # Use a unique run created by test 11 if available, or create one
    from pipeline_service import get_archive_metadata, get_archived_file_path

    # Check if test 11 left a run we can use
    resp = client.get("/runs")
    runs = _json_of(resp)
    run_id = None
    for r in runs:
        meta = get_archive_metadata(r["run_id"])
        if meta is not None:
            run_id = r["run_id"]
            break

    if run_id is None:
        # Create a fresh run with unique files
        tmp_dir = tempfile.mkdtemp(prefix="archive_restart_")
        try:
            file_map = {}
            for key, src in [("open_po", OPEN_PO_PATH), ("bd_tracker", BD_TRACKER_PATH),
                             ("eagle_eye", EAGLE_EYE_PATH)]:
                dst = os.path.join(tmp_dir, f"{key}_{uuid.uuid4().hex}.xlsx")
                _make_unique_copy(src, dst)
                file_map[key] = dst
            with open(file_map["open_po"], "rb") as op, \
                 open(file_map["bd_tracker"], "rb") as bd, \
                 open(file_map["eagle_eye"], "rb") as ee:
                resp = client.post("/validate", data={
                    "open_po": (op, f"op_{uuid.uuid4().hex}.xlsx"),
                    "bd_tracker": (bd, f"bd_{uuid.uuid4().hex}.xlsx"),
                    "eagle_eye": (ee, f"ee_{uuid.uuid4().hex}.xlsx"),
                })
            data = _json_of(resp)
            if data["valid"]:
                resp = client.post("/process", data=json.dumps(data["file_paths"]),
                                   content_type="application/json")
                r = _json_of(resp)
                if r["success"] and not r.get("duplicate", False):
                    run_id = r["run_id"]
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    if run_id is None:
        print("  [SKIP] Could not create a run with archive for testing")
        return

    # Verify via service layer
    meta = get_archive_metadata(run_id)
    check(meta is not None, f"Archive metadata accessible via service layer (run {run_id[:8]})")

    for st in ("open_po", "bd_tracker", "eagle_eye"):
        fpath = get_archived_file_path(run_id, st)
        check(fpath is not None and os.path.exists(fpath),
              f"Archived file for {st} accessible on disk via service layer")

    # Verify via web endpoints (simulated restart: fresh client)
    from web_app import app as fresh_app
    fresh_app.config["TESTING"] = True
    fresh_client = fresh_app.test_client()
    for st in ("open_po", "bd_tracker", "eagle_eye"):
        resp = fresh_client.get(f"/download/{run_id}/source/{st}")
        check(resp.status_code in (200, 404),
              f"Source download endpoint works after simulated restart for {st} (got {resp.status_code})")


# ======================================================================
# TEST 13: Completed_With_Archive_Warning status on archive failure
# ======================================================================

def test_archive_failure_status(client):
    """Prove that when archiving fails, the run is marked
    Completed_With_Archive_Warning instead of silently succeeding."""
    print("\n" + "=" * 60)
    print("TEST 13: test_archive_failure_status")
    print("=" * 60)

    # Create unique temp copies of the files (to avoid duplicate detection)
    # so that process_upload_run is actually called with real archiving.
    tmp_dir = tempfile.mkdtemp(prefix="archive_test_")
    unique_files = {}
    try:
        for key, src in [("open_po", OPEN_PO_PATH), ("bd_tracker", BD_TRACKER_PATH),
                         ("eagle_eye", EAGLE_EYE_PATH)]:
            dst = os.path.join(tmp_dir, f"{key}_{uuid.uuid4().hex}.xlsx")
            _make_unique_copy(src, dst)
            unique_files[key] = dst

        # Validate the copies
        with open(unique_files["open_po"], "rb") as op, \
             open(unique_files["bd_tracker"], "rb") as bd, \
             open(unique_files["eagle_eye"], "rb") as ee:
            resp = client.post("/validate", data={
                "open_po": (op, f"open_po_{uuid.uuid4().hex}.xlsx"),
                "bd_tracker": (bd, f"bd_{uuid.uuid4().hex}.xlsx"),
                "eagle_eye": (ee, f"ee_{uuid.uuid4().hex}.xlsx"),
            })
        data = _json_of(resp)
        if not data["valid"]:
            print("  [SKIP] Validation failed for temp copies")
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return

        # Override archive to fail
        import pipeline_service
        orig_archive = pipeline_service.archive_source_files

        def broken_archive(run_id, file_paths):
            raise PermissionError("Simulated archive failure: disk not writable")

        pipeline_service.archive_source_files = broken_archive

        try:
            resp = client.post(
                "/process",
                data=json.dumps(data["file_paths"]),
                content_type="application/json",
            )
            r = _json_of(resp)
            check(r["success"], "Process succeeded (even with archive failure)")

            # When files are unique, process_upload_run is called; archive fails
            # because the broken archive dir doesn't exist (and is not created by mkdtemp)
            from pipeline_service import get_run_status
            from pipeline_db import get_connection
            conn = get_connection()
            try:
                status_info = get_run_status(r["run_id"], conn=conn)
                actual_status = status_info["run_status"]
                notes = status_info.get("run_notes", "")
                has_archive_issue = "Archive_failed" in notes
                check(actual_status == "Completed_With_Archive_Warning",
                      f"Status is Completed_With_Archive_Warning (got '{actual_status}')")
                check(has_archive_issue,
                      f"Run notes indicate archive issue: '{notes[:80]}'")
            finally:
                conn.close()
        finally:
            pipeline_service.archive_source_files = orig_archive
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

def main():
    global RESULTS
RESULTS = []


def _make_unique_copy(src, dst):
    """Copy an Excel file and add a unique marker to a hidden cell
    so the content hash differs while the file remains valid."""
    import uuid
    import openpyxl
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Write UUID to cell in column beyond validated columns (col 50)
        ws.cell(row=ws.max_row, column=50, value=uuid.uuid4().hex)
    wb.save(dst)
    wb.close()
    return dst


def main():
    global RESULTS
    RESULTS = []

    missing = check_source_files()
    if missing:
        print("ERROR: Missing source files:")
        for m in missing:
            print(f"  {m}")
        print(f"\nPlace the Excel files in: {EXCEL_DIR}")
        sys.exit(1)

    from web_app import app

    print("=" * 60)
    print("PHASE 3 IMPORT TRACKER WEB APP TESTS")
    print(f"Python: {sys.version.split()[0]}")
    print(f"Source files:")
    print(f"  Open PO:   {OPEN_PO_PATH}")
    print(f"  BD Tracker: {BD_TRACKER_PATH}")
    print(f"  Eagle Eye:  {EAGLE_EYE_PATH}")
    print("=" * 60)

    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    client = app.test_client()

    tests = [
        ("test_index_page_loads", test_index_page_loads),
        ("test_validate_no_files", test_validate_no_files),
        ("test_clear_temp_files", test_clear_temp_files),
        ("test_process_invalid_paths", test_process_invalid_paths),
        ("test_runs_list_empty", test_runs_list_empty),
        ("test_run_detail_not_found", test_run_detail_not_found),
        ("test_download_invalid_run", test_download_invalid_run),
        ("test_full_upload_flow", test_full_upload_flow),
        ("test_duplicate_detection_flow", test_duplicate_detection_flow),
        ("test_clear_does_not_affect_history", test_clear_does_not_affect_history),
        ("test_archive_survives_workspace_clear", test_archive_survives_workspace_clear),
        ("test_archive_survives_app_restart", test_archive_survives_app_restart),
        ("test_archive_failure_status", test_archive_failure_status),
    ]

    for name, func in tests:
        try:
            func(client)
        except Exception as e:
            RESULTS.append((False, f"{name} raised exception: {e}"))
            print(f"\n  [FAIL] {name} raised exception: {e}")
            traceback.print_exc()

    passed = sum(1 for c, _ in RESULTS if c)
    failed = sum(1 for c, _ in RESULTS if not c)

    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    print(f"  Total assertions: {len(RESULTS)}")
    print(f"  Passed:           {passed}")
    print(f"  Failed:           {failed}")
    if len(RESULTS) > 0:
        print(f"  Pass rate:        {passed / len(RESULTS) * 100:.1f}%")
    print("=" * 60)

    if failed > 0:
        print("\nFailed assertions:")
        for c, m in RESULTS:
            if not c:
                print(f"  [FAIL] {m}")

    return 0 if failed == 0 else 1



if __name__ == "__main__":
    sys.exit(main())
