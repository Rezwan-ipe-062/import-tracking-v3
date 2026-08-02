"""Anchor - Streamlit UI for the Planning Team import-visibility pipeline.

Runs on the manager's laptop. Accepts the four source Excel files (Mode A) or a
pre-generated Import Visibility Master workbook (Mode B), calls the existing
Phase-3 scripts (via pipeline.py) unchanged, and renders master + risk +
data-quality views. Local persistence lives under .anchor/ (gitignored).
"""

import datetime
import io

import pandas as pd
import streamlit as st

import pipeline
import logic
import store
import theme
from ui import components as C
from ui.components import (brand_block, section, topbar, empty_state,
                           info_note, kpi_row)

# Spec v3.1.0 - exactly six nav pages.
PAGES = [
    "Action Centre",
    "PO Journey",
    "Shipment Visibility",
    "Risk & Exposure",
    "Data Quality",
    "Thresholds & Refresh",
]

UPLOAD_LABELS = [
    ("open", "Open PO", "Defines the active population of POs."),
    ("tracker", "BD Tracker", "Import milestones: LC, SI, RDD, ETA/ETD, OBL, final docs."),
    ("ee", "Eagle Eye", "Container & shipment visibility: From, DDPO, container, tracking."),
    ("threshold", "Country Thresholds", "Agreed timing rules per route (optional)."),
]

FRESH_DAYS = 3


def _css():
    st.markdown(theme.inject_css(), unsafe_allow_html=True)


def _goto(view):
    st.session_state["anchor_view"] = view
    st.rerun()


def _active_df(ctx):
    headers = ctx.get("master_headers") or []
    rows = ctx.get("master") or []
    df = C.to_pandas(headers, rows)
    if df.empty:
        return df
    if "Population Status" in df.columns:
        df = df[df["Population Status"].astype(str).str.strip().str.lower() == "active"]
    return df


def _fmt_q(v):
    try:
        if v is None:
            return "-"
        return f"{float(v):,.0f}"
    except (TypeError, ValueError):
        return "-" if v in ("", None) else str(v)


def _is_empty(v):
    if v is None:
        return True
    if isinstance(v, float) and v != v:
        return True
    return isinstance(v, str) and not v.strip()


def _n_missing(master, headers, col):
    if col not in headers:
        return None
    i = headers.index(col)
    return sum(1 for r in master if i < len(r) and _is_empty(r[i]))


# --------------------------------------------------------------------------- #
# Welcome (no local view yet)
# --------------------------------------------------------------------------- #

def welcome_empty():
    _css()
    brand_block()
    st.markdown(f'<div style="max-width:560px;margin:6vh auto;text-align:center">'
                f'{theme.logo_mark(64)}'
                f'<div style="font-size:2.2rem;font-weight:700;margin:.7rem 0 .2rem">Anchor</div>'
                f'<div class="muted" style="font-size:1.02rem">Import visibility and '
                f'action prioritisation</div>'
                f'<div class="muted" style="margin:1.2rem auto">Upload the latest source '
                f'files to create a refreshed import action view.</div>', unsafe_allow_html=True)
    if st.button("Upload Latest Files", type="primary"):
        _goto("upload")
    st.markdown('<div class="muted" style="text-align:center">Use Open PO, BD Tracker, '
                'Eagle Eye and Country Threshold files from the same refresh cycle '
                'where possible.</div>')


# --------------------------------------------------------------------------- #
# Restore (a local view exists)
# --------------------------------------------------------------------------- #

def restore_screen():
    _css()
    brand_block()
    meta = store.load_view_meta()
    state, note = pipeline.freshness_state(meta)
    st.markdown("## Restore or refresh")
    st.markdown(f'<div class="restore-banner"><div><b>Locally restored view</b> - '
                f'verify freshness before acting. {note}</div></div>', unsafe_allow_html=True)
    kpi_row([
        ("Open POs", meta.get("open_po_count", "-"), "plain"),
        ("Pipeline", "v" + (meta.get("version") or "-"), "plain"),
        ("Last refreshed", str(meta.get("refreshed_at", "-"))[:16], "plain"),
    ])
    c1, c2, c3 = st.columns(3)
    if c1.button("Restore Latest Dashboard", type="primary", width="stretch"):
        ctx = pipeline.build_context_from_disk()
        st.session_state["context"] = ctx
        st.session_state["page"] = "Action Centre"
        _goto("app")
    if c2.button("Upload New Full Set", width="stretch"):
        _goto("upload")
    if c3.button("Clear Local Data & Start Fresh", width="stretch"):
        st.session_state["confirm_clear"] = True
        st.rerun()

    if st.session_state.get("confirm_clear"):
        _clear_dialog()


# --------------------------------------------------------------------------- #
# Upload + validation + processing
# --------------------------------------------------------------------------- #

def upload_screen():
    _css()
    brand_block()
    st.markdown("## Upload the latest source data")
    mode = st.radio("Upload mode", ["A - Full source refresh (four files)",
                                    "B - Pre-generated Master workbook"],
                    horizontal=True, key="upload_mode")
    if mode and mode.startswith("B"):
        _upload_mode_b()
    else:
        _upload_mode_a()


def _upload_mode_a():
    info_note("Four files, ideally from the same refresh cycle. "
              "No data leaves this device.")
    ups = {}
    for key, label, hint in UPLOAD_LABELS:
        ups[key] = st.file_uploader(label, type=["xlsx", "xlsm", "xlsb"],
                                    key=f"up_{key}", help=hint)

    section("Validation")
    checks = _validation_checks(ups)
    _render_checks(checks)

    if not ups.get("threshold"):
        info_note("Country Thresholds is optional - Anchor will use built-in "
                  "defaults if it is not provided.")

    ready = checks[0]["ok"] and checks[1]["ok"] and checks[2]["ok"]
    if st.button("Generate Anchor Action View", type="primary", disabled=not ready):
        with st.spinner("Cleaning and classifying the four files..."):
            try:
                _run_mode_a(ups)
            except Exception as exc:
                st.error(f"Could not generate the view. {exc}")
                st.info("Check that the files are the expected Anchor source files "
                        "and try again.")


def _upload_mode_b():
    info_note("Upload a pre-generated Import Visibility Master workbook. Anchor "
              "reads it as-is; open quantities are shown at PO level only.")
    up = st.file_uploader("Import Visibility Master", type=["xlsx", "xlsm", "xlsb"],
                          key="up_master")
    if up is not None:
        st.markdown(f'- **{up.name}** - {_human_size(up.size)} - will be validated '
                    'when you generate.')
    if st.button("Generate Anchor Action View", type="primary", disabled=up is None):
        with st.spinner("Reading the master workbook..."):
            try:
                _run_mode_b(up)
            except Exception as exc:
                st.error(f"Validation failed. {exc}")
                st.info("The master workbook was not in the expected Anchor layout. "
                        "Check the required columns and retry.")


def _human_size(n):
    if n >= 1 << 20:
        return f"{n / (1 << 20):.1f} MB"
    if n >= 1 << 10:
        return f"{n / (1 << 10):.0f} KB"
    return f"{n} B"


def _validation_checks(ups):
    checks = []
    for key, label, _h in UPLOAD_LABELS:
        checks.append({"ok": ups.get(key) is not None,
                       "required": key != "threshold",
                       "label": f"{label} file recognised"})
    return checks


def _render_checks(checks):
    for c in checks:
        icon = "✓" if c["ok"] else "·"
        required = " (required)" if c["required"] else " (optional)"
        st.markdown(f"- `{icon}` {c['label']}{required}")


def _run_mode_a(ups):
    staged, meta = {}, {}
    for key in ("open", "tracker", "ee", "threshold"):
        u = ups.get(key)
        if u is not None:
            staged[key], meta[key] = pipeline.stage_upload(u, key)
        else:
            staged[key], meta[key] = None, None
    context = pipeline.run(staged, staged.get("threshold"))
    store.save_view(context)
    st.session_state["context"] = context
    st.session_state["page"] = "Action Centre"
    _goto("app")


def _run_mode_b(upload):
    """Read a pre-built master workbook into a view context.

    The workbook is already the source of truth: no cleaning/merge is performed.
    Every non-empty row is treated as an active master row.
    """
    import openpyxl
    data = upload.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name.lower() in ("master", "import visibility master"):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The master workbook is empty.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers = [h for h in headers if h]
    master = []
    for r in rows[1:]:
        if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        row = list(r)[:len(headers)]
        while len(row) < len(headers):
            row.append(None)
        master.append(row)

    if "Urgency" not in headers:
        headers = headers + ["Urgency"]
        for r in master:
            r.append("Monitor")

    refresh = datetime.datetime.now().replace(microsecond=0)
    meta = {
        "version": "user-master",
        "refreshed_at": refresh,
        "threshold_filename": "import visibility master",
        "threshold_version": "import visibility master",
        "open_po_count": len(master),
        "source_files": [{"key": "master", "filename": upload.name,
                          "loaded_at": refresh.isoformat()}],
        "master_headers": headers,
        "bd_headers": [], "ee_headers": [], "op_headers": [],
        "is_restored": False,
    }
    context = {
        "master": master, "master_headers": headers,
        "control": {}, "bd_rows": [], "ee_rows": [], "op_rows": [],
        "summary": {"po_count": len(master), "severity": {}}, "meta": meta,
        "is_restored": False,
    }
    store.save_view(context)
    st.session_state["context"] = context
    st.session_state["page"] = "Action Centre"
    _goto("app")


# --------------------------------------------------------------------------- #
# App shell
# --------------------------------------------------------------------------- #

def render_app():
    ctx = st.session_state.get("context") or pipeline.build_context_from_disk()
    if not ctx.get("master"):
        st.session_state["anchor_view"] = "restore" if store.has_view() else "welcome"
        st.rerun()

    st.session_state["context"] = ctx

    page = st.sidebar.radio("View", PAGES,
                            index=_page_index(), key="nav_radio")
    st.session_state["page"] = page

    if page != "Thresholds & Refresh":
        topbar(ctx.get("meta", {}))

    if page == "Action Centre":
        action_centre(ctx)
    elif page == "PO Journey":
        po_journey(ctx)
    elif page == "Shipment Visibility":
        shipment_visibility(ctx)
    elif page == "Risk & Exposure":
        risk_and_exposure(ctx)
    elif page == "Data Quality":
        data_quality(ctx)
    else:
        thresholds_and_refresh(ctx)


def _page_index():
    prev = st.session_state.get("page", "Action Centre")
    return list(PAGES).index(prev) if prev in PAGES else 0


# --------------------------------------------------------------------------- #
# Global search
# --------------------------------------------------------------------------- #

def _global_search(ctx, key_prefix="gs"):
    headers = ctx.get("master_headers") or []
    rows = ctx.get("master") or []
    po_idx = headers.index("Purchasing Document") if "Purchasing Document" in headers else None
    if po_idx is None:
        return None
    fields_idx = []
    for col in ("Purchasing Document", "Short Text", "Material (AGI)", "Container No."):
        if col in headers:
            fields_idx.append(headers.index(col))

    q = st.text_input("Global search", placeholder="PO, product / AGI, or container",
                      key=f"{key_prefix}_q").strip()
    if not q:
        return None
    q_l = q.lower()
    cands = set()
    for r in rows:
        blob = " | ".join("" if i >= len(r) else str(r[i]) for i in fields_idx)
        if q_l in blob.lower() and po_idx < len(r):
            cands.add(str(r[po_idx]))
    cands = sorted(cands)
    if cands:
        sel = st.selectbox("Search results", ["—"] + cands, key=f"{key_prefix}_pick")
        if st.button("Open PO journey", key=f"{key_prefix}_open"):
            if sel and sel != "—":
                st.session_state["po"] = sel
                st.session_state["page"] = "PO Journey"
                st.rerun()
    else:
        st.caption("No PO, product, AGI or container matched.")
    return None


# --------------------------------------------------------------------------- #
# Action Centre
# --------------------------------------------------------------------------- #

def action_centre(ctx):
    df = _active_df(ctx)
    headers = ctx.get("master_headers") or []
    if df.empty:
        empty_state("No open actions", "All POs are complete or monitor-only.")
        return

    sev = df["Urgency"].value_counts().to_dict() if "Urgency" in df.columns else {}
    open_count = int(df["Purchasing Document"].nunique()) \
        if "Purchasing Document" in df.columns else len(df)
    no_bd = int(df["BD Tracker ETA"].isna().sum()) if "BD Tracker ETA" in df.columns else 0
    no_ee = int(df["EE ETA"].isna().sum()) if "EE ETA" in df.columns else 0

    kpi_row([
        ("Active Open POs", open_count, "plain"),
        ("Critical", sev.get("Critical", 0), "crit"),
        ("Urgent", sev.get("Urgent", 0), "urg"),
        ("Data Review", sev.get("Data Review", 0), "dr"),
        ("Monitor", sev.get("Monitor", 0), "mon"),
        ("No BD record", no_bd, "plain"),
        ("No EE evidence", no_ee, "plain"),
    ])

    section("Open requirement by unit", "KG and litre ('L') units are shown "
            "separately - never combined.")
    qty = logic.qty_by_unit(ctx.get("master") or [], headers, population_only=True)
    if qty:
        cols = st.columns(len(qty))
        for i, (unit, v) in enumerate(qty.items()):
            with cols[i]:
                st.markdown(
                    f'<div class="card" style="margin:.15rem 0">'
                    f'<div class="kpi-value">{_fmt_q(v)}</div>'
                    f'<div class="kpi-label">Open PO requirement &middot; {unit}</div></div>',
                    unsafe_allow_html=True)

    section("Priority actions", "Sorted Critical > Urgent > Data Review > Monitor, "
            "then RDD ascending.")
    pdf = _priority_table(df, headers)
    if pdf.empty:
        empty_state("No open actions", "All POs are complete or monitor-only.")
    else:
        st.dataframe(pdf, width="stretch", height=520, hide_index=True)

    section("Open a PO journey")
    polist = sorted({str(p) for p in df["Purchasing Document"].dropna().unique()})
    sel = st.selectbox("Drill into a PO", [""] + polist, key="ac_po")
    if sel:
        st.session_state["po"] = sel
        st.session_state["page"] = "PO Journey"
        st.rerun()


def _priority_table(df, headers):
    keep = [c for c in ["Purchasing Document", "Urgency", "Primary Reason",
                        "Still to be Delivered (Qty)", "Order Unit", "RDD",
                        "Overall Status", "Import Country", "Supplier Name",
                        "Short Text"] if c in df.columns]
    out = df[keep].copy()

    if "Primary Reason" in out.columns:
        prox = out["Primary Reason"].fillna("").astype(str)
        out["Required follow-up"] = [logic.suggested_followup(x)[0] for x in prox]
        out["Suggested owner *"] = [logic.suggested_followup(x)[1] for x in prox]

    if "Urgency" in out.columns:
        conf = []
        for _, row in df.iterrows():
            conf.append(logic.data_confidence(list(row), list(df.columns)))
        out["Confidence"] = conf

        out["_sort"] = out["Urgency"].map(C.priority_sort_key)
        out["_no_rdd"] = out["RDD"].isna() if "RDD" in out.columns else 0
        out = out.sort_values(["_sort", "_no_rdd", "RDD"], na_position="last")
        out = out.drop(columns=["_sort", "_no_rdd"])
        if "Purchasing Document" in out.columns:
            out = out.drop_duplicates(subset=["Purchasing Document"])
    return out


# --------------------------------------------------------------------------- #
# PO Journey
# --------------------------------------------------------------------------- #

def po_journey(ctx):
    df = _active_df(ctx)
    headers = ctx.get("master_headers") or []
    po_col = "Purchasing Document"
    if po_col not in df.columns or df.empty:
        empty_state("No PO data", "Restore or upload a view first.")
        return

    c_search, c_po = st.columns([2.4, 1])
    with c_search:
        _global_search(ctx, "pj")
    with c_po:
        polist = sorted({str(p) for p in df[po_col].dropna().unique()})
        sel = st.session_state.get("po")
        idx = polist.index(str(sel)) if sel and str(sel) in polist else 0
        po = st.selectbox("PO", polist, index=idx)
        st.session_state["po"] = po

    sub = df[df[po_col].astype(str) == str(po)]
    if sub.empty:
        empty_state("PO not found", "This PO is not in the current view.")
        return
    one = sub.iloc[0]

    ug = str(one.get("Urgency", "Monitor"))
    kind = ("crit" if ug == "Critical" else "urg" if ug == "Urgent"
            else "dr" if ug == "Data Review" else "mon")
    kpi_row([
        ("PO", str(po), "plain"),
        ("Urgency", ug, kind),
        ("Open Qty", _fmt_q(one.get("Still to be Delivered (Qty)")), "plain"),
        ("RDD", C.fmt_date(one.get("RDD")), "plain"),
        ("Import Country", str(one.get("Import Country") or "-"), "plain"),
        ("Confidence", confidence_for(one, list(df.columns)), "plain"),
    ])

    section("Milestone journey")
    rows = []
    for col, lbl, note in (
            ("LC Date", "LC", "LC issued"),
            ("SI Shared Date", "SI", "SI shared"),
            ("ETD", "ETD", "Schedule / ETD"),
            ("BD Tracker ETA", "BD ETA", "BD expected arrival"),
            ("EE ETA", "EE ETA", "Shipment ETA"),
            ("OBL/EBL rcvd Date", "OBL", "OBL/EBL received"),
            ("Final Docs rcvd Date", "Final", "Final docs received")):
        v = one.get(col)
        done = isinstance(v, (datetime.date, datetime.datetime)) and not pd.isna(v)
        rows.append([lbl,
                     C.fmt_date(v) if done else "-",
                     "Complete" if done else "Missing",
                     note if done else f"{note}: not recorded"])
    st.dataframe(pd.DataFrame(rows, columns=["Milestone", "Date", "Status", "Note"]),
                 width="stretch", hide_index=True)

    section("Partial shipments", "Open quantity is at PO level - partial rows are "
            "process detail and are not totalled as shipment quantity.")
    cols_p = [c for c in ["Partial Shipment No.", "Overall Status", "ETD", "EE ETA",
                          "OBL/EBL rcvd Date", "Final Docs rcvd Date"] if c in df.columns]
    if cols_p:
        st.dataframe(sub[cols_p], width="stretch", hide_index=True)
    st.markdown('<div class="muted">Open quantity relates to the PO. Partial-shipment '
                'rows describe process steps and must not be totalled.</div>')

    if "Container No." in df.columns and sub["Container No."].isna().all():
        st.markdown('<div class="restore-banner"><div><b>Container evidence not '
                    'confirmed.</b> The link between this PO and container records is '
                    'not recorded; verify before closing the risk.</div></div>',
                    unsafe_allow_html=True)

    section("Manager follow-up")
    reason = str(one.get("Primary Reason") or "")
    fup, owner = logic.suggested_followup(reason)
    st.markdown(f"- **Primary reason:** {reason}")
    st.markdown(f"- **Suggested follow-up:** {fup}")
    st.markdown(f"- **Suggested owner (suggested):** {owner}")
    st.markdown('<div class="muted">Follow-up and owner are derived from Primary '
                'Reason and are recommendations, not assignments.</div>')

    notes = store.load_notes()
    note = notes.get(str(po), "")
    st.text_area("Note (saved on this device only)", value=note, key=f"note_{po}")
    if st.button("Save note"):
        notes[str(po)] = st.session_state[f"note_{po}"]
        store.save_notes(notes)
        st.success("Note saved.")


def confidence_for(one, headers):
    try:
        return logic.data_confidence(list(one), headers)
    except Exception:
        return "Medium"


# --------------------------------------------------------------------------- #
# Shipment Visibility
# --------------------------------------------------------------------------- #

def shipment_visibility(ctx):
    df = _active_df(ctx)
    if df.empty:
        empty_state("No shipment data", "Restore or upload a view first.")
        return
    st.markdown('<div class="muted">One row per container / evidence record. Open '
                'quantity is never summed here (PO level only).</div>')
    keep = [c for c in ["Purchasing Document", "From", "Container No.", "Tracking",
                        "Status", "EE ETD", "EE ETA", "Import Country"] if c in df.columns]
    view = df[keep].drop_duplicates()
    st.dataframe(view, width="stretch", height=520, hide_index=True)

    section("Shipment evidence")
    def has(c):
        return int(df[c].notna().sum()) if c in df.columns else 0
    kpi_row([
        ("Container assigned", has("Container No."), "plain"),
        ("with EE ETA", has("EE ETA"), "plain"),
        ("with EE ETD", has("EE ETD"), "plain"),
    ])


# --------------------------------------------------------------------------- #
# Risk & Exposure
# --------------------------------------------------------------------------- #

def risk_and_exposure(ctx):
    df = _active_df(ctx)
    if df.empty:
        empty_state("No risk data", "Restore or upload a view first.")
        return
    st.markdown("Counts are distinct POs, never per-row across partial shipments; "
                "no KG/L quantity is combined.")
    crit = df[df["Urgency"].isin(["Critical", "Urgent"])] \
        if "Urgency" in df.columns else df

    c1, c2 = st.columns(2)
    with c1:
        section("Critical / Urgent by Import country")
        if "Import Country" in crit.columns and "Purchasing Document" in crit.columns:
            s = crit.groupby("Import Country")["Purchasing Document"].nunique().sort_values(ascending=False)
            st.bar_chart(s)
        else:
            st.caption("No country field in this view.")
    with c2:
        section("Critical / Urgent by Supplier")
        if "Supplier Name" in crit.columns and "Purchasing Document" in crit.columns:
            s = crit.groupby("Supplier Name")["Purchasing Document"].nunique().sort_values(ascending=False).head(12)
            st.bar_chart(s)
        else:
            st.caption("No supplier field in this view.")

    section("Risk reason x country matrix")
    if {"Import Country", "Primary Reason", "Urgency", "Purchasing Document"} <= set(df.columns):
        sub = df[df["Urgency"].isin(["Critical", "Urgent"])]
        tab = sub.pivot_table(index="Import Country", columns="Primary Reason",
                              values="Purchasing Document", aggfunc="nunique", fill_value=0)
        st.dataframe(tab, width="stretch", height=260)
    else:
        st.caption("Matrix unavailable - required fields not present.")

    section("Product exposure")
    if "Short Text" in df.columns and "Purchasing Document" in df.columns:
        grp = df.groupby("Short Text").agg(
            Open_PO=("Purchasing Document", "nunique"),
            Qty=("Still to be Delivered (Qty)", "sum")
            if "Still to be Delivered (Qty)" in df.columns else ("Purchasing Document", "count"),
            Crit=("Urgency", lambda s: int(s.eq("Critical").sum())),
            Urg=("Urgency", lambda s: int(s.eq("Urgent").sum())),
            DR=("Urgency", lambda s: int(s.eq("Data Review").sum())),
            Earliest_RDD=("RDD", "min") if "RDD" in df.columns else ("Purchasing Document", "count"),
        ).reset_index().rename(columns={"Short Text": "Product / AGI"})
        st.dataframe(grp, width="stretch", hide_index=True)
        st.markdown("<div class='muted'>Quantities carry their labelled unit and are "
                    "never cross-unit summed.</div>", unsafe_allow_html=True)
    else:
        st.caption("No product field in this view.")

    section("RDD exposure horizon")
    if "RDD" in df.columns:
        def bucket(row):
            v = row["RDD"]
            off = logic.rdd_offset(v) if not pd.isna(v) else None
            return logic.rdd_horizon(off) if off is not None else "Unknown"
        h = df.assign(horizon=df.apply(bucket, axis=1)) \
                .groupby("horizon")["Purchasing Document"].nunique()
        h = h.reindex(["Overdue", "0-7d", "8-30d", "31-60d", ">60d", "Unknown"]).fillna(0)
        st.bar_chart(h)
    else:
        st.caption("No RDD field in this view.")

    info_note("Exposure summaries for awareness only; use Data Quality exports for "
              "a controlled copy.")


# ---------------------------------------------------------------------------
# Data Quality
# ---------------------------------------------------------------------------

def data_quality(ctx):
    headers = ctx.get("master_headers") or []
    master = ctx.get("master") or []
    control = ctx.get("control") or {}

    def missing(col):
        if col not in headers:
            return "-"
        i = headers.index(col)
        return sum(1 for r in master if i < len(r) and _is_empty(r[i]))

    q_rows = [
        ("RDD missing", missing("RDD")),
        ("Route / country unknown", missing("Import Country")),
        ("No BD record", missing("BD Tracker ETA")),
        ("No Eagle Eye record", missing("EE ETA")),
        ("BD PO not in Open PO", len(ctx.get("bd_rows") or [])),
        ("EE PO not in Open PO", len(ctx.get("ee_rows") or [])),
        ("Status complete but open qty", missing("Overall Status")),
        ("Container not assigned", missing("Container No.")),
    ]
    section("Reconciliation & quality KPI")
    kpi_row([(label, v, "plain") for label, v in q_rows])
    st.markdown('<div class="legend">Data-review flagged rows use the slate Data '
                'Review label; a data gap is not treated as Critical.</div>')

    section("Exception queue")
    names = [n for n in ("Exceptions", "Unmatched BD", "Unmatched EE",
                         "Cleaning Log") if control.get(n)]
    if names:
        tabs = st.tabs(names)
        for name, tab in zip(names, tabs):
            with tab:
                _sheet(control.get(name), name)
    else:
        info_note("No control sheets in this view.")

    if "Urgency" in headers:
        i = headers.index("Urgency")
        dr = sum(1 for r in master if str(r[i]).strip() == "Data Review")
        st.markdown(f'<div class="restore-banner"><div><b>{dr}</b> row(s) with '
                    'Urgency = Data Review (slate, not red).</div></div>',
                    unsafe_allow_html=True)


def _sheet(item, name):
    if not item:
        empty_state("No data", f"No {name} entries in this view.")
        return
    h, rows, _ = item
    st.dataframe(C.to_pandas(h, rows), width="stretch", hide_index=True)


# ---------------------------------------------------------------------------
# Thresholds & Refresh
# ---------------------------------------------------------------------------

def thresholds_and_refresh(ctx):
    brand_block()
    meta = ctx.get("meta", {})
    import rule_engine

    section("Thresholds")
    th = ctx.get("thresholds") or rule_engine.DEFAULT_THRESHOLDS
    info_note(f'Active route threshold file: <b>{meta.get("threshold_filename") or "built-in defaults"}</b>')
    rows = []
    for kind, label in (("LC", "LC"), ("ETD", "Schedule / ETD")):
        for route in ("India", "ASEAN", "ChinaEA", "Europe"):
            urgent, critical = th.get(kind, {}).get(route, (0, 0))
            rows.append([route, label, urgent, critical,
                         f"If '{label}' is missing {urgent}d before RDD it is Urgent; "
                         f"{critical}d it is Critical.", "-", "-"])
    st.dataframe(pd.DataFrame(rows, columns=["Import route", "Milestone",
                                             "Urgent", "Critical", "Rule",
                                             "Effective date", "Updated"]),
                 width="stretch", hide_index=True)
    info_note("Threshold editing is read-only here; update the Country Thresholds "
              "sheet and re-upload to change timing.")

    section("Source file freshness")
    _freshness_table(meta)

    section("Local data & refresh")
    info_note(f'Current view: <b>{"restored" if meta.get("is_restored") else "fresh"}</b>'
              f' - last refreshed {str(meta.get("refreshed_at", "-"))[:16]}.')
    kpi_row([("Notes saved", store.notes_count(), "plain"),
             ("Open POs", meta.get("open_po_count", "-"), "plain")])

    c1, c2, c3 = st.columns(3)
    if c1.button("Upload New Full Set", width="stretch"):
        _goto("upload")
    if c2.button("Export Current Master", width="stretch"):
        _export_master(ctx)
    if c3.button("Clear Local Data & Start Fresh", width="stretch"):
        st.session_state["confirm_clear"] = True
        st.rerun()

    _export_panel(ctx)

    st.markdown('<div class="muted">Source files (last run):</div>')
    for s in meta.get("source_files", []):
        st.markdown(f'- **{s.get("filename", "-")}**')

    if st.session_state.get("confirm_clear"):
        _clear_dialog()


def _freshness_table(meta):
    rows = []
    for s in meta.get("source_files", []):
        name = s.get("filename", "")
        if not name:
            continue
        loaded = s.get("loaded_at", "")
        rows.append([name, str(loaded)[:16], FRESH_DAYS, _state_of_loaded(loaded)])
    if rows:
        st.dataframe(pd.DataFrame(rows, columns=["Source file", "Loaded at",
                                                 "Fresh window (days)", "Status"]),
                     width="stretch", hide_index=True)


def _state_of_loaded(loaded):
    if not loaded:
        return "unknown"
    try:
        dt = datetime.datetime.fromisoformat(str(loaded))
        return "FRESH" if (datetime.datetime.now() - dt).days <= FRESH_DAYS else "STALE"
    except ValueError:
        return "unknown"


def _export_master(ctx):
    try:
        path = _write_excel_master(ctx)
        st.success(f"Exported {path}")
    except Exception as exc:
        st.error(f"Export failed: {exc}")


def _write_excel_master(ctx):
    out = store.export_dir() / "anchor_master.xlsx"
    headers = ctx.get("master_headers") or []
    rows = ctx.get("master") or []
    import pandas as pd
    out.write_bytes(b"")
    pd.DataFrame(rows, columns=headers).to_excel(out, index=False)
    return out


def _export_panel(ctx):
    headers = ctx.get("master_headers") or []
    master = ctx.get("master") or []
    df = (_active_df(ctx))
    meta = ctx.get("meta", {})
    if not headers:
        return
    n1, n2, n3, n4 = st.columns(4)
    if n1.button("Export action list", help="Controlled copy of the priority queue"):
        _export(ctx, df, "anchor_actions.csv", "Priority actions",
                "Current priority actions")
    if n2.button("Export PO journey"):
        _export(ctx, master, "anchor_po_journey.csv", "PO journey detail",
                "PO journey rows for this view")
    if n3.button("Export Data Quality"):
        _export(ctx, master, "anchor_reconciliation.csv", "Data Quality / Reconciliation",
                "Master rows with quality reconciliation notes")
    if n4.button("Export Shipment"):
        ship = [c for c in ["Purchasing Document", "From", "Container No.",
                            "EE ETA", "Status"] if c in headers]
        ship_rows = [[r[headers.index(c)] for c in ship] for r in master]
        _export_raw(ship, ship_rows, "anchor_shipments.csv", "Shipment evidence",
                    "One row per container / evidence record")


def _export_raw(header, rows, filename, subject, filter_desc):
    meta = st.session_state.get("context", {}).get("meta", {})
    try:
        path = C.export_csv(header, rows, filename, meta, filter_desc, subject)
        st.success(f"Exported {path}")
    except Exception as exc:
        st.error(f"Export failed: {exc}")


def _export(ctx, rows, filename, subject, filter_desc):
    meta = ctx.get("meta", {})
    headers = ctx.get("master_headers") or []
    if isinstance(rows, pd.DataFrame):
        record = [list(r) for r in rows.values.tolist()]
    else:
        record = rows
    try:
        path = C.export_csv(headers, record, filename, meta, filter_desc, subject)
        st.success(f"Exported {path}")
    except Exception as exc:
        st.error(f"Export failed: {exc}")


def _clear_dialog():
    st.markdown("---")
    st.markdown("**Clear all Anchor data stored on this device?** This removes the "
                "master, exceptions, filters, notes, source-file metadata and the "
                "temporary working set. The original Excel files on your device or "
                "in OneDrive are <b>not</b> deleted.")
    if st.session_state.get("confirm_clear_again"):
        c1, c2 = st.columns(2)
        if c2.button("Confirm clear all", type="primary"):
            store.clear_all(confirmed=True)
            st.session_state.pop("confirm_clear", None)
            st.session_state.pop("confirm_clear_again", None)
            st.session_state.pop("context", None)
            st.session_state.pop("page", None)
            _goto("welcome")
    else:
        c1, c2, c3 = st.columns(3)
        if c1.button("Cancel", width="stretch"):
            st.session_state["confirm_clear"] = False
            st.rerun()
        if c3.button("I understand, clear everything", type="primary",
                     width="stretch"):
            st.session_state["confirm_clear_again"] = True
            st.rerun()


# --------------------------------------------------------------------------- #
# Dispatch
# --------------------------------------------------------------------------- #

def main():
    _css()
    view = st.session_state.get("anchor_view")
    if view is None:
        st.session_state["anchor_view"] = "welcome" if not store.has_view() else "restore"
        st.rerun()
    elif view == "welcome":
        welcome_empty()
    elif view == "restore":
        restore_screen()
    elif view == "upload":
        upload_screen()
    elif view == "app":
        render_app()
    else:
        st.session_state["anchor_view"] = "welcome" if not store.has_view() else "restore"
        st.rerun()


if __name__ == "__main__":
    main()