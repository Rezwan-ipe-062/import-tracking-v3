"""Single-file cleaner for the Eagle Eye import file.

Run this script in VS Code (Run button / F5):

    1. A file dialog opens - pick the Eagle Eye .xlsx file.
    2. The script reads the data sheet, keeps only the agreed fields
       (From / DDPO prefix / cleansed DDPO / Container No. / Tracking /
       Status / ETA / ETD), splits the F/G prefix off the DDPO number, and
       normalises every DDPO to a 10-digit digit string so it can be joined
       with the other import files later (the merge step does the joining).
    3. It writes <name>_cleaned.xlsx next to the source file (the source
       file is never modified) and prints a summary to the console.

Kept rows stay at container / tracking-line level - one PO can have many
container rows, and each stays a separate row.

ETA and ETD are written as real Excel dates (yyyy-mm-dd) when present, never
as text. The '-' placeholders in the source become blank cells. DDPO is
written as the prefix letter plus the bare 10-digit number
(e.g. G6590028423 -> 'G', '6590028423').

Requires: openpyxl  (pip install openpyxl)
"""

import datetime
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REQUIRED = {
    "from",
    "ddpo",
    "containerno",
    "tracking",
    "status",
    "eta",
    "etd",
}

OUT_COLUMNS = [
    "From",
    "Prefix (F/G)",
    "Cleansed DDPO",
    "Container No.",
    "Tracking",
    "Status",
    "ETA",
    "ETD",
]

# Statuses observed in the Eagle Eye data. Used only to flag unexpected values.
KNOWN_STATUSES = {
    "1 Pending TP Flag / CCR",
    "2 To be booked",
    "3 Booked",
    "4 Sailed",
    "5 Arrived at Port",
    "6 Arrived at Door",
}

DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%b-%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%Y.%m.%d",
)


def norm(s):
    return "".join(c for c in str(s).lower() if c.isalnum())


def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()


def is_blankish(v):
    if v is None:
        return True
    if isinstance(v, (int, float)) and v == 0:
        return True
    return str(v).strip().lower() in ("", "-", "--", "na", "n/a", "n.a.",
                                      "nan", "null", "none")


def po_text(v):
    """Return a PO cell as plain digit text.

    Excel often stores PO numbers as numbers, and when a file is re-saved or
    edited the values silently become floats (6590028423.0). Strip the .0 so
    the 10-digit match works instead of dropping the row.
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return clean_str(v)


def find_header_row(sheet):
    for r in range(1, min(sheet.max_row, 6) + 1):
        names = set()
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(row=r, column=c).value
            if v is not None:
                names.add(norm(v))
        if {"from", "ddpo", "status"} <= names:
            return r
    return None


def build_column_map(sheet, header_row):
    colmap = {}
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=header_row, column=c).value
        if v is not None:
            colmap[norm(v)] = c
    missing = REQUIRED - set(colmap)
    return colmap, missing


def live_row(a, c, i):
    return not (is_blankish(a) and is_blankish(c) and is_blankish(i))


def iter_indexed_rows(ws, colmap, header_row):
    """Yield (excel_row_number, (From, DDPO, Status)) for the data region.

    Lazy scan with a dead-run break (50 consecutive empty rows) and a
    peek-ahead safety net (200 rows), so the table extent is discovered, not
    hard-coded.
    """
    ca, cc, ci = (colmap["from"], colmap["ddpo"], colmap["status"])
    r = header_row + 1
    dead_run = 0
    while r <= ws.max_row:
        a = ws.cell(row=r, column=ca).value
        c = ws.cell(row=r, column=cc).value
        i = ws.cell(row=r, column=ci).value
        if live_row(a, c, i):
            dead_run = 0
            yield r, (a, c, i)
            r += 1
            continue
        dead_run += 1
        r += 1
        if dead_run < 50:
            continue
        # Suspected end - peek ahead before giving up.
        resumed = False
        for rr in range(r, min(r + 200, ws.max_row + 1)):
            a2 = ws.cell(row=rr, column=ca).value
            c2 = ws.cell(row=rr, column=cc).value
            i2 = ws.cell(row=rr, column=ci).value
            if live_row(a2, c2, i2):
                resumed = True
                break
        if not resumed:
            break
        dead_run = 0


def parse_ddpo(raw):
    """Return (prefix, cleansed_ddpo).

    Accepts 'G6590028423' -> ('G', '6590028423'),
    'F6590028423' -> ('F', '6590028423').
    Unparsable values come back as ('', raw) so they can be flagged.
    """
    s = po_text(raw)
    m = re.match(r"^([A-Za-z])(\d{10})$", s)
    if m:
        return m.group(1).upper(), m.group(2)
    return "", s


def parse_date_value(v):
    """Return (datetime_or_None, error_or_None). Time is stripped."""
    if v is None:
        return None, None
    if isinstance(v, datetime.datetime):
        return v.replace(hour=0, minute=0, second=0, microsecond=0), None
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day), None
    if isinstance(v, (int, float)):
        if v == 0:
            return None, None
        try:
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
            return d.replace(hour=0, minute=0, second=0, microsecond=0), None
        except (OverflowError, ValueError, TypeError):
            return None, "unreadable date value %r" % (v,)
    s = str(v).strip()
    if is_blankish(v):
        return None, None
    for fmt in DATE_FORMATS:
        try:
            d = datetime.datetime.strptime(s, fmt)
            return d.replace(hour=0, minute=0, second=0, microsecond=0), None
        except ValueError:
            continue
    return None, "invalid date: %r" % s


def cell_value(fcell, vcell):
    """Prefer the raw (formula) cell; fall back to the computed value."""
    if isinstance(fcell, str) and fcell.startswith("="):
        return vcell if vcell is not None else fcell
    return fcell


def clean_to_rows(src_path):
    """Clean the Eagle Eye file in memory and return (headers, rows, info).

    ``rows`` is the list of kept output rows (same shape as OUT_COLUMNS).
    ``info`` carries the diagnostics needed for the console summary and is
    also used by clean_merge.py.
    """
    wb_formulas = openpyxl.load_workbook(src_path, data_only=False)
    wb_values = openpyxl.load_workbook(src_path, data_only=True)
    sheet_name = wb_formulas.active.title
    ws = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]

    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError('Could not find a header row containing "From" / "DDPO" / "Status".')
    colmap, missing = build_column_map(ws, header_row)
    if missing:
        found = [ws.cell(row=header_row, column=c).value
                 for c in range(1, ws.max_column + 1)]
        raise ValueError("Missing expected columns: %s\nHeader found: %s"
                         % (", ".join(sorted(missing)), found))

    kept = []
    dropped = []
    review = []
    raw_ddpos = []
    n_blank_rows = 0
    kept_pos = []

    for r, (a, c, i) in iter_indexed_rows(ws, colmap, header_row):
        if is_blankish(a) and is_blankish(c) and is_blankish(i):
            n_blank_rows += 1
            continue

        src_from = clean_str(cell_value(ws.cell(row=r, column=colmap["from"]).value,
                                        ws_v.cell(row=r, column=colmap["from"]).value))
        ddpo_raw = po_text(cell_value(ws.cell(row=r, column=colmap["ddpo"]).value,
                                      ws_v.cell(row=r, column=colmap["ddpo"]).value))
        container = clean_str(cell_value(ws.cell(row=r, column=colmap["containerno"]).value,
                                         ws_v.cell(row=r, column=colmap["containerno"]).value))
        tracking = clean_str(cell_value(ws.cell(row=r, column=colmap["tracking"]).value,
                                        ws_v.cell(row=r, column=colmap["tracking"]).value))
        status = clean_str(cell_value(ws.cell(row=r, column=colmap["status"]).value,
                                      ws_v.cell(row=r, column=colmap["status"]).value))

        prefix, ddpo = parse_ddpo(ddpo_raw)

        if ddpo == "":
            dropped.append([r, ddpo_raw, "Blank / unparsable DDPO number"])
            continue
        if prefix == "":
            dropped.append([r, ddpo_raw, "Unparsable DDPO format"])
            continue

        flags = []
        if src_from == "":
            flags.append("Blank From")
        if status == "":
            flags.append("Blank Status")
        elif status not in KNOWN_STATUSES:
            flags.append("Unrecognised Status: %r" % status)

        eta_raw = cell_value(ws.cell(row=r, column=colmap["eta"]).value,
                             ws_v.cell(row=r, column=colmap["eta"]).value)
        eta, err = parse_date_value(eta_raw)
        if err:
            flags.append("ETA: %s" % err)

        etd_raw = cell_value(ws.cell(row=r, column=colmap["etd"]).value,
                             ws_v.cell(row=r, column=colmap["etd"]).value)
        etd, err2 = parse_date_value(etd_raw)
        if err2:
            flags.append("ETD: %s" % err2)

        kept.append([src_from, prefix, ddpo, container, tracking, status, eta, etd])
        review.append([r, ddpo, flags])
        kept_pos.append(ddpo)
        raw_ddpos.append(ddpo_raw)

    info = {
        "src_path": src_path,
        "sheet_name": sheet_name,
        "kept": kept,
        "review": review,
        "dropped": dropped,
        "kept_pos": kept_pos,
        "raw_ddpos": raw_ddpos,
        "n_blank_rows": n_blank_rows,
    }
    return OUT_COLUMNS, kept, info


def write_rows(dst_path, headers, rows, date_cols=(), title="Cleaned Output",
               width=24, date_format="yyyy-mm-dd"):
    """Write headers + rows to a fresh workbook with basic styling."""
    out = openpyxl.Workbook()
    ws_out = out.active
    ws_out.title = title
    for c, name in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True)
    for i, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_out.cell(row=i, column=c, value=val)
            if c in date_cols and isinstance(val, datetime.datetime):
                cell.number_format = date_format
    for c in range(1, len(headers) + 1):
        ws_out.column_dimensions[ws_out.cell(row=1, column=c).column_letter].width = width
    ws_out.freeze_panes = "A2"
    out.save(dst_path)


def clean_workbook(src_path, dst_path):
    headers, kept, info = clean_to_rows(src_path)

    write_rows(dst_path, headers, kept, date_cols={7, 8},
               title="Cleaned Eagle Eye", width=24)

    dropped = info["dropped"]
    review = info["review"]
    kept_pos = info["kept_pos"]
    n_kept = len(kept)
    n_dropped = len(dropped)
    n_distinct = len(set(kept_pos))
    n_flagged = sum(1 for row in review if row[2])
    print("=" * 60)
    print("Eagle Eye cleaner - summary")
    print("=" * 60)
    print("Source          : %s" % src_path)
    print("Sheet           : %s" % info["sheet_name"])
    print("Kept            : %d rows across %d distinct PO numbers" % (n_kept, n_distinct))
    print("Dropped         : %d  (blank DDPO: %d)"
          % (n_dropped,
             sum(1 for d in dropped if "Blank" in d[2])))
    print("Blank rows skip : %d" % info["n_blank_rows"])
    print("Rows flagged    : %d" % n_flagged)
    print("Output          : %s" % dst_path)
    print("=" * 60)
    if n_flagged:
        print("Flagged rows to review (Source Row):")
        for row in review:
            if row[2]:
                print("  row %s PO=%s -> %s" % (row[0], row[1], "; ".join(row[2])))
    if n_dropped:
        print("Dropped rows (Source Row):")
        for d in dropped:
            print("  row %s PO=%r -> %s" % (d[0], d[1], d[2]))


def pick_file(title):
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    root.destroy()
    if path:
        return Path(path)
    p = input("No file chosen. Type the full path to the .xlsx: ").strip()
    return Path(p) if p else None


def main():
    src = pick_file("Select the Eagle Eye .xlsx file")
    if src is None:
        print("No file selected - nothing to do. Exiting.")
        return
    if not src.exists():
        print("File not found: %s" % src)
        return
    dst = src.with_name(src.stem + "_cleaned.xlsx")
    try:
        clean_workbook(src, dst)
    except ValueError as exc:
        print("Error:", exc)


if __name__ == "__main__":
    main()
